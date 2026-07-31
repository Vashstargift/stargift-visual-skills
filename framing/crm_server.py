#!/usr/bin/env python3
"""StarGift CRM MCP server for the staff Doc.

Exposes ONE tool: my_deals(tg_user_id, limit).
Role scoping is enforced server-side by crm-deals.php:
  - manager / staff  → own deals only
  - owner / admin / director → all deals

The identity plugin (stargift_identity) injects the tg_user_id into context
so the LLM always passes the correct value without guessing.
"""
import sys

sys.path.insert(0, "/Users/docbrown/hermes-doc")

from mcp.server.fastmcp import FastMCP
from crm_client import call
import selection_sender
from selection_sender import _fetch_products, _fmt_price

mcp = FastMCP("stargift-crm")


@mcp.tool()
def my_deals(tg_user_id: str, limit: int = 20) -> str:
    """СОБСТВЕННЫЕ сделки сотрудника (закреплённые за ним), независимо от роли.

    Менеджеру сервер и так отдаёт только его сделки; но админу/директору/owner он
    отдаёт ВСЕ сделки CRM — поэтому здесь дополнительно оставляем только сделки,
    закреплённые за самим написавшим (его «мои»). Для просмотра всех сделок —
    это не тот инструмент.
    tg_user_id — Telegram id написавшего (берётся из контекста сессии).
    limit — максимальное количество строк в ответе (по умолчанию 20).
    """
    if not tg_user_id:
        return "Не указан tg_user_id — не могу определить сотрудника."

    try:
        data = call("crm-deals.php", tg=str(tg_user_id))
    except Exception as e:
        return f"Ошибка доступа к сделкам: {e}"

    # crm-deals.php returns a bare JSON array; guard against object wrapper too.
    deals = data if isinstance(data, list) else data.get("deals", [])

    # "Мои" = закреплённые за написавшим. Админам сервер отдаёт все сделки, поэтому
    # сужаем по имени сотрудника здесь (резолвим его роль/имя из tg).
    try:
        who = call("bot-resolve-user.php", tg=str(tg_user_id), params={"tg": str(tg_user_id)})
        my_name = (who or {}).get("name", "") if isinstance(who, dict) else ""
    except Exception:
        my_name = ""
    if my_name:
        deals = [d for d in deals if (d.get("assigned_to") or "") == my_name]

    if not deals:
        who_txt = f" за {my_name}" if my_name else ""
        return f"Сделок{who_txt} не найдено."

    # Resolve client name + contact for the shown deals (cache lookups by client_id).
    client_cache = {}

    def _client_brief(cid):
        if not cid:
            return ""
        if cid not in client_cache:
            name, contact = "", ""
            try:
                c = call("crm-clients.php", tg=str(tg_user_id), params={"id": str(cid)})
                if isinstance(c, dict):
                    name = (c.get("full_name")
                            or " ".join(p for p in (c.get("first_name"), c.get("last_name")) if p)
                            or c.get("company") or "")
                    for key, pref in (("phone", ""), ("whatsapp", "WA "), ("telegram", "@"), ("email", "")):
                        v = c.get(key)
                        if v:
                            contact = pref + str(v)
                            break
            except Exception:
                pass
            client_cache[cid] = f"{name} ({contact})" if name and contact else (name or contact)
        return client_cache[cid]

    # Newest first (server may already sort; we cap to limit regardless).
    lines = []
    for d in deals[:limit]:
        amt = d.get("amount")
        if amt is not None:
            try:
                amt_s = f"{int(float(amt)):,} ₽".replace(",", " ")
            except (ValueError, TypeError):
                amt_s = str(amt)
        else:
            amt_s = "—"

        title = d.get("title") or "—"
        client = _client_brief(d.get("client_id")) or "клиент не указан"
        lines.append(
            f"• {client} — {title} — {d.get('status', '')} — {amt_s} (#{d.get('id')})"
        )

    total = len(deals)
    shown = min(limit, total)
    return f"Твоих сделок: {total} (показаны {shown}):\n" + "\n".join(lines)


@mcp.tool()
def deal_edit(tg_user_id: str, deal_id: str, changes: dict, confirm: bool = False) -> str:
    """Правка сделки. БЕЗ confirm=true — только ПРЕДПРОСМОТР (что изменится), запись НЕ происходит.
    С confirm=true — применяет изменение.

    deal_id — ТОЛЬКО точный id из листинга (в скобках после #, напр. 8974f3abd98f3701).
    НИКОГДА не передавай заголовок/название сделки — сервер такое отклонит (404).
    Если id не знаешь — сначала возьми его из my_deals / deals_at_risk.

    Сервер сам проверяет права: менеджер/staff может править только СВОИ сделки
    (иначе откажет), директор/owner — любые. changes — словарь полей, напр.
    {"status": "won"} или {"notes": "..."}.
    tg_user_id — Telegram id написавшего (из контекста сессии).
    """
    if not tg_user_id:
        return "Не указан tg_user_id — не могу определить сотрудника."
    if not deal_id or not isinstance(changes, dict) or not changes:
        return "Нужны deal_id и непустой changes (что менять)."
    import re as _re
    deal_id = str(deal_id).strip().lstrip("#")
    if not _re.fullmatch(r"[0-9a-fA-F]{16}|[0-9a-fA-F\-]{36}", deal_id):
        return (f"«{deal_id}» — это не id сделки. Нужен точный id из листинга "
                f"(в скобках после #, напр. 8974f3abd98f3701) — возьми его из my_deals / deals_at_risk.")
    if not confirm:
        pretty = ", ".join(f"{k} → {v}" for k, v in changes.items())
        return (f"ПРЕДПРОСМОТР: сделка #{deal_id} изменится так: {pretty}.\n"
                f"Если всё верно — подтверди, и я применю (вызову deal_edit с confirm=true).")
    body = dict(changes)
    body["id"] = deal_id
    try:
        call("crm-deals.php", tg=str(tg_user_id), method="PATCH", body=body)
    except Exception as e:
        # 403 (чужая сделка / нет прав) тоже попадёт сюда
        return f"Не удалось изменить сделку #{deal_id}: {e} (возможно, нет прав на эту сделку)."
    return f"Готово — сделка #{deal_id} обновлена ({', '.join(changes)})."


def _find_products(query: str, limit: int = 3) -> list:
    """Поиск экспоната для генераторов (ценник/сертификат/пачка): точная фраза,
    при пустой выдаче — fallback по словам со скорингом (как в catalog_search).
    «Тайсон перчатка» находит «Майк Тайсон — Перчатка с автографом»."""
    try:
        found = _fetch_products(query, limit, 0)
    except Exception:
        return []
    if found:
        return found
    words = [w.strip(",.«»\"") for w in query.split() if len(w.strip(",.«»\"")) >= 3]
    words.sort(key=lambda w: (not w[:1].isupper(), -len(w)))
    pool, seen = [], set()
    for w in words[:4]:
        try:
            extra = _fetch_products(w, limit, 0)
        except Exception:
            continue
        for pr in extra:
            pid = str(pr.get("id") or "")
            if pid and pid not in seen:
                seen.add(pid)
                pool.append(pr)
    if not pool:
        return []
    ql = [w.lower() for w in words]

    def _score(pr):
        hay = ((pr.get("title") or "") + " " + (pr.get("description") or "")).lower()
        return sum(1 for w in ql if w in hay)

    pool.sort(key=_score, reverse=True)
    return pool[:limit]


def _catalog_filter(products: list, min_price: int, in_stock_only: bool,
                    gallery: str = "") -> list:
    """Фильтры выдачи каталога: нижняя граница цены и «только в наличии».
    Позиции без цены при min_price НЕ отбрасываются (цена по запросу может
    быть любой) — они и так ранжируются в хвост. Пометка total_matched
    (сколько подходило всего) прокидывается сквозь фильтр."""
    out = []
    for p in products:
        price = float(p.get("price") or 0)
        if min_price > 0 and price > 0 and price < min_price:
            continue
        if in_stock_only:
            av = str(p.get("availability") or "").strip()
            if av in ("preorder", "outofstock", "0"):
                continue
        if gallery and gallery.lower() not in (p.get("gallery") or "").lower():
            continue
        out.append(p)

    class _ProductList(list):
        total_matched = 0

    dropped = len(products) - len(out)
    r = _ProductList(out)
    r.total_matched = max(getattr(products, "total_matched", len(products)) - dropped, len(out))
    return r


@mcp.tool()
def catalog_search(query: str = "", limit: int = 8, max_price: int = 0,
                   min_price: int = 0, in_stock_only: bool = False,
                   category: str = "", gallery: str = "") -> str:
    """Поиск ЭКСПОНАТОВ в каталоге StarGift + их НАЛИЧИЕ (галерея / под заказ).
    Это ЕДИНСТВЕННЫЙ источник наличия экспонатов (автографы, фото, футболки…) —
    вопросы «что в наличии по X», «есть ли автограф Y» идут СЮДА, а не в склад
    расходников supply_stock (там только акрил/подставки/упаковка).
    query — имя/тема (напр. 'Месси', 'хоккей'). Для нескольких персон — ОТДЕЛЬНЫЙ
    вызов на каждую. limit — сколько позиций (по умолч. 8). max_price — потолок ₽.
    min_price — НИЖНЯЯ граница ₽: клиент сказал «от 400 тысяч» → min_price=400000
    (дешёвые позиции исключаются — предлагать кубок за 30К при брифе «от 400К» нельзя).
    in_stock_only=True — ТОЛЬКО реально в наличии (галереи): менеджер сказал
    «в наличии», «что стоит в галереях» → под заказ и частные коллекции исключаются.
    Если просят «в наличии, а если нет — под заказ» — два вызова: сперва с
    in_stock_only=True, потом без.
    category — ТОЧНОЕ имя категории каталога, лучший инструмент для брифов «кому»:
    Женщине / Мужчине / Руководителю / Бизнесмену / Ребенку / Музыка / Рок /
    Поп-музыка / Кино / Актеры / Спорт / Футбол / Бокс и ММА / История /
    Россия (общий) / Великие деятели России / Искусство / Книги / Другие страны.
    «Подборка для женщины 50+» → category='Женщине' (можно без query!);
    «руководителю» → category='Руководителю'. Комбинируй с query и ценами.
    gallery — фильтр по КОНКРЕТНОЙ галерее («Времена года» / «Гименей» /
    «Dream House»): менеджер сказал «что стоит во Временах года» → gallery=…
    (подразумевает наличие — комбинируй с in_stock_only=True).
    Возвращает: название, тип, цена, НАЛИЧИЕ (галерея/под заказ), ссылка, фото.
    """
    # Uses paginated mode (?page=1) — returns {products:[...], total:N, ...}.
    # Legacy mode (no page= param) ignores search and serves cached full catalog.
    #
    # Verified column mapping from catalog-compat.php legacy positional format (lines 292-300)
    # for reference (paginated mode returns named fields; mapping kept as comment):
    #   [0] product_id  [1] null  [2] null  [3] label  [4] categories (;-sep)
    #   [5] person/name  [6] description/type  [7] photo_url (space-sep)
    #   [8] price (str)  [9] availability  [10-12] null  [13] uid  [14] parent_uid
    #   [15-19] null  [20] shortMeta  [21] tags

    if not query and not category:
        return "Нужен query или category."
    try:
        products = _catalog_filter(_fetch_products(query, limit, max_price, category=category),
                                   min_price, in_stock_only, gallery)
    except Exception as e:
        return f"Ошибка при обращении к каталогу: {e}"

    # Длинная фраза («Месси футболка сборной Аргентины») часто не матчится целиком,
    # хотя по фамилии всё находится — при пустой выдаче пробуем слова по отдельности.
    # ВСЕ слова (не только первое удачное): раньше «Лев Яшин» останавливался на
    # «Лев» и отдавал Льва Ошанина. Теперь кандидаты со всех слов, скоринг по числу
    # совпавших слов запроса — «Яшин» (2 слова) выигрывает у «Ошанина» (1 слово).
    fallback_note = ""
    if not products:
        words = [w.strip(",.«»\"") for w in query.split() if len(w.strip(",.«»\"")) >= 3]
        words.sort(key=lambda w: (not w[:1].isupper(), -len(w)))
        seen_ids = set()
        for w in words[:4]:
            try:
                extra = _catalog_filter(_fetch_products(w, limit, max_price), min_price, in_stock_only, gallery)
            except Exception:
                continue
            for pr in extra:
                pid = str(pr.get("id") or "")
                if pid and pid not in seen_ids:
                    seen_ids.add(pid)
                    products.append(pr)
        if products:
            qwords = [w.lower() for w in words]
            def _match_score(pr):
                hay = ((pr.get("title") or "") + " " + (pr.get("description") or "")).lower()
                return sum(1 for w in qwords if w in hay)
            products.sort(key=_match_score, reverse=True)
            best = _match_score(products[0])
            # Слабые хвосты (совпало на 2+ слова меньше лучшего) отрезаем
            products = [p for p in products if _match_score(p) >= max(1, best - 1)][:limit]
            fallback_note = (f"Точной фразы «{query}» не нашёл — показываю ближайшее "
                             f"по словам запроса (проверь глазами):\n\n")

    try:
        if not products:
            return f"По запросу «{query}» в каталоге ничего не нашёл."

        blocks = []
        for p in products:
            try:
                product_id = p.get("id") or ""
                name = (p.get("title") or "").strip()
                ptype = (p.get("description") or "").strip()
                photos = p.get("photos") or []
                first_photo = photos[0] if photos else ""
                price_str = _fmt_price(p.get("price"))
                url_product = f"https://stargift.ru/product/{product_id}/"

                # Наличие: из карточки каталога (availability+gallery), НЕ со склада расходников
                av = str(p.get("availability") or "").strip()
                gal = (p.get("gallery") or "").strip()
                if av == "preorder" or av == "outofstock":
                    stock_line = "Наличие: под заказ"
                elif av == "0":
                    stock_line = "Наличие: частная коллекция (не продаётся)"
                elif gal:
                    stock_line = f"Наличие: в галерее «{gal}»"
                else:
                    stock_line = "Наличие: в наличии (галерея не указана)"
                block = f"«{name}» — {ptype}\nЦена: {price_str}\n{stock_line}\nСсылка: {url_product}"
                if first_photo:
                    block += f"\nФото: {first_photo}"
                blocks.append(block)
            except Exception:
                continue

        if not blocks:
            return f"По запросу «{query}» в каталоге ничего не нашёл."

        header = f"Нашёл {len(blocks)} вариантов:"
        # Честность о полноте (Вашик, 22.07): менеджер должен знать, всё ли это.
        total = getattr(products, "total_matched", len(blocks))
        if len(blocks) < limit and total <= len(blocks):
            footer = f"\n\nЭто ВСЕ подходящие позиции ({len(blocks)}) — скажи менеджеру честно, что больше нет."
        elif total > len(blocks):
            footer = (f"\n\nПо теме есть ещё ~{total - len(blocks)}+ подходящих экспонатов — "
                      f"предложи менеджеру прислать ещё.")
        else:
            footer = ""
        return fallback_note + header + "\n\n" + "\n\n".join(blocks) + footer

    except Exception as e:
        return f"Ошибка при обработке данных каталога: {e}"


@mcp.tool()
def send_selection(tg_user_id: str, query: str = "", limit: int = 8, max_price: int = 0,
                   min_price: int = 0, in_stock_only: bool = False,
                   category: str = "", product_ids: list = None) -> str:
    """Отправить клиенту/в чат ПОДБОРКУ товаров как фото-альбом (реальные картинки с подписями) в Telegram.
    Используй, когда нужна подборка С ФОТО. tg_user_id — Telegram id чата (из контекста сессии).
    query — тема/имя. limit — сколько карточек (базово 8, макс 10). max_price — потолок цены (0 = без).
    min_price — нижняя граница ₽ (бриф «от 400 тысяч» → 400000, дешёвое не попадёт).
    in_stock_only=True — только реально в наличии в галереях («в наличии» от менеджера
    = ЭТОТ флаг; под заказ в такую подборку не попадает).
    category — категория каталога (список в catalog_search) для брифов «кому».
    product_ids — КУРАТОРСКИЙ режим (лучший для сложных брифов: «женщине 50+»,
    «зарубежная музыка», «врачу»): сначала собери кандидатов через catalog_search
    (по категориям/персонам), ОТБЕРИ ГОЛОВОЙ подходящие по смыслу (ты знаешь, кто
    зарубежный, кто советский, что уместно врачу или женщине), и передай сюда
    список product_id в нужном порядке — уйдут ровно они.
    Подпись каждого фото: «Название» — тип / цена / ссылка. После отправки верни краткое
    подтверждение. Результат тула содержит пометку о полноте («это все» / «есть ещё») —
    ОБЯЗАТЕЛЬНО передай её менеджеру своими словами: подходит мало — честно скажи сколько;
    есть ещё — предложи «на эту тему есть ещё экспонаты, прислать?».
    """
    if not tg_user_id:
        return "Не указан tg_user_id."
    if not query and not category and not product_ids:
        return "Нужен query, category или product_ids."
    return selection_sender.send_album(tg_user_id, query, limit, max_price, min_price,
                                       in_stock_only, category, product_ids)


@mcp.tool()
def create_reminder(tg_user_id: str, title: str, remind_at: str, client_name: str = "") -> str:
    """Создать напоминание сотруднику (синкается в его Яндекс-календарь + push).
    title — текст напоминания (напр. «связаться с клиентом Ивановым»).
    remind_at — дата и время в формате 'YYYY-MM-DD HH:MM:SS'. Вычисли из «завтра/через N дней/в пятницу»
    сам, опираясь на текущее время из контекста сессии.
    client_name — имя клиента (необязательно). tg_user_id — Telegram id из контекста.
    """
    if not tg_user_id or not title or not remind_at:
        return "Нужны title и remind_at в формате 'YYYY-MM-DD HH:MM:SS'."
    try:
        who = call("bot-resolve-user.php", tg=str(tg_user_id), params={"tg": str(tg_user_id)})
    except Exception as e:
        return f"Не удалось определить сотрудника: {e}"
    if not who.get("linked"):
        return "Твой Telegram не привязан к CRM — напоминание не создать. Привяжи аккаунт через /link."
    body = {"manager": who.get("name", ""), "title": title, "remind_at": remind_at}
    if client_name:
        body["client_name"] = client_name
    try:
        call("crm-reminders.php", tg=str(tg_user_id), method="POST", body=body)
    except Exception as e:
        return f"Не удалось создать напоминание: {e}"
    return f"Напоминание создано: «{title}» на {remind_at} (синкнул в твой Яндекс-календарь)."


@mcp.tool()
def catalog_get(product_id: str) -> str:
    """Полная карточка товара по id (из catalog_search): название, краткое и
    ПОЛНОЕ описание, цена, размеры, категории, фото. Используй перед правкой
    текстов (catalog_edit), чтобы видеть текущее полное описание."""
    if not (product_id or "").strip():
        return "Нужен product_id."
    import urllib.parse as _up
    import urllib.request as _ur
    url = ("https://stargift.ru/api/catalog-compat.php?page=1&ids="
           + _up.quote(str(product_id).strip()))
    try:
        import json as _json
        data = _json.loads(_ur.urlopen(url, timeout=30).read())
    except Exception as e:
        return f"Ошибка каталога: {e}"
    prods = data.get("products", [])
    if not prods:
        return f"Карточка {product_id} не найдена."
    p = prods[0]
    dims = p.get("dimensions")
    lines = [
        f"«{p.get('title')}» (#{p.get('id')})",
        f"Краткое: {p.get('description') or '—'}",
        f"Цена: {_fmt_price(p.get('price'))}",
        f"Категории: {', '.join(p.get('categories') or [])}",
        f"Размеры: {dims if dims else '—'}",
        f"Фото ({len(p.get('photos') or [])}): " + " ".join(p.get("photos") or []),
        "ПОЛНОЕ ОПИСАНИЕ:",
        (p.get("fullDescription") or "—"),
    ]
    return "\n".join(lines)


@mcp.tool()
def catalog_edit(tg_user_id: str, product_query: str, changes: dict, confirm: bool = False) -> str:
    """Правка карточки товара (цена/описание/метка и т.п.). БЕЗ confirm=true — ПРЕДПРОСМОТР (запись не идёт).
    product_query — название товара для поиска. changes — словарь полей: price, description, label, gallery,
    sort_order, availability и т.п. (длинный текст 'text' идёт на согласование старшему).
    Сервер сам решает по роли: опубликовать сразу или отправить на согласование. Менеджер без прав — откажет.
    tg_user_id — Telegram id из контекста.
    """
    if not tg_user_id:
        return "Не указан tg_user_id."
    if not product_query or not isinstance(changes, dict) or not changes:
        return "Нужны product_query (какой товар) и changes (что менять)."
    q = str(product_query).strip().lstrip("#")
    # Менеджеры часто присылают ССЫЛКУ на карточку или id со слагом
    # (https://stargift.ru/product/sg_…_hex-lionel-messi/) — извлекаем чистый id,
    # иначе запрос уходил в поиск по названию и «карточка не найдена».
    import re as _re
    _m = _re.search(r"(sg_\d+_[a-f0-9]+)", q) or _re.search(r"/product/(\d+)", q)
    if _m:
        q = _m.group(1)
    # Прямое попадание по id (sg_… или цифровой id из catalog_get/catalog_search)
    if q.startswith("sg_") or q.isdigit():
        import urllib.parse as _up
        import urllib.request as _ur
        import json as _json
        try:
            data = _json.loads(_ur.urlopen(
                "https://stargift.ru/api/catalog-compat.php?page=1&ids=" + _up.quote(q),
                timeout=30).read())
            prods = data.get("products", [])
        except Exception as e:
            return f"Ошибка каталога: {e}"
        if not prods:
            return f"Карточка с id {q} не найдена."
    else:
        try:
            prods = _fetch_products(product_query, 5, 0)
        except Exception as e:
            return f"Ошибка поиска товара: {e}"
    if not prods:
        return f"Товар по запросу «{product_query}» не найден."
    if len(prods) > 1:
        opts = "\n".join(f"• {p.get('title')} (#{p.get('id')}) — {_fmt_price(p.get('price'))}" for p in prods[:5])
        return ("Нашёл несколько товаров — уточни, какой именно (назови точнее или укажи #id):\n" + opts)
    p = prods[0]
    pid, pname = p.get("id"), p.get("title")
    if not confirm:
        pretty = ", ".join(f"{k} → {v}" for k, v in changes.items())
        return (f"ПРЕДПРОСМОТР: у «{pname}» (#{pid}) изменится: {pretty}.\n"
                f"Подтверди — применю (catalog_edit с confirm=true).")
    body = dict(changes)
    # Полное описание на сервере — поле `text` (texts-таблица + синк в
    # full_description + сброс кэша). Модели часто шлют full_description —
    # алиасим, иначе сервер МОЛЧА игнорирует неизвестное поле и отвечает ok.
    for alias in ("full_description", "fullDescription", "полное_описание"):
        if alias in body and "text" not in body:
            body["text"] = body.pop(alias)
    _ALLOWED = {"person", "description", "price", "price_old", "gallery",
                "dim_length", "dim_width", "dim_height", "weight", "label",
                "sort_order", "availability", "categories", "cert_number",
                "recipients", "occasions", "tags", "fulfillment_meta", "text"}
    unknown = [k for k in body if k not in _ALLOWED]
    if unknown and not (set(body) - set(unknown)):
        return (f"Поля {unknown} сервер не принимает — ничего не изменилось бы. "
                f"Доступные: description (краткое), text (ПОЛНОЕ описание), price, "
                f"label, gallery, availability, categories и др.")
    body["product_id"] = pid
    try:
        res = call("catalog-edit.php", tg=str(tg_user_id), method="POST", body=body)
    except Exception as e:
        return f"Не удалось изменить карточку: {e} (возможно, нет прав)."
    if isinstance(res, dict) and res.get("queued"):
        return f"Правка «{pname}» отправлена на согласование старшему менеджеру."
    if isinstance(res, dict) and res.get("ok"):
        return f"Готово — карточка «{pname}» обновлена ({', '.join(changes)})."
    return f"Ответ сервера: {res}"


@mcp.tool()
def send_cards(tg_user_id: str, product_names: list) -> str:
    """Отправить ОТОБРАННЫЕ тобой товары подборкой — каждый отдельным фото-сообщением.
    product_names — список ТОЧНЫХ названий экспонатов, которые ты сам нашёл через catalog_search
    и счёл подходящими под тему. Используй для подборки по абстрактной/узкой теме: сначала несколько
    раз вызови catalog_search по смежным словам, отбери релевантные, потом пришли их сюда.
    tg_user_id — Telegram id из контекста.
    """
    if not tg_user_id:
        return "Не указан tg_user_id."
    if not product_names or not isinstance(product_names, list):
        return "Передай product_names — список точных названий товаров."
    import selection_sender
    return selection_sender.send_cards_by_names(str(tg_user_id), product_names)


@mcp.tool()
def deal_create(tg_user_id: str, title: str, client: str = "", amount: int = 0,
                status: str = "new", notes: str = "", product_id: str = "",
                sale_date: str = "", confirm: bool = False) -> str:
    """Создать НОВУЮ сделку в CRM за написавшим менеджером. БЕЗ confirm=true — ПРЕДПРОСМОТР
    (ничего не записывается); с confirm=true — создаёт. ВСЕГДА показывай предпросмотр и
    жди «да» менеджера перед confirm=true.

    title — суть сделки, напр. «Футболка Зидана с автографом» (если продаётся конкретный
      экспонат — назови его; product_id из catalog_search привяжет его к сделке).
    client — имя/телефон клиента для поиска в CRM. Если клиент не найден или найдено
      несколько — предпросмотр покажет варианты; сделку можно создать и без клиента.
    amount — сумма в рублях. status — new | in_progress | won | lost (по умолчанию new).
    sale_date — ДАТА ПРОДАЖИ 'YYYY-MM-DD'. Для won-продаж прошлых дней ОБЯЗАТЕЛЬНА:
      менеджер назвал дату («от 25.06») → передай её; назвал только месяц («продажи мая»)
      → последний день месяца и добавь в notes «дата примерная, день не указан»;
      даты нет вовсе → СПРОСИ менеджера, не заноси молча сегодняшним днём
      (инцидент 22.07: 30 сделок легли одним днём и исказили месяцы).
    notes — свободный комментарий менеджера. tg_user_id — из контекста сессии.
    """
    if not tg_user_id:
        return "Не указан tg_user_id — не могу определить сотрудника."
    if not (title or "").strip():
        return "Нужен title — что продаём."

    # Кто создаёт — имя менеджера (сделка закрепляется за ним)
    try:
        who = call("bot-resolve-user.php", tg=str(tg_user_id), params={"tg": str(tg_user_id)})
        manager = (who or {}).get("name", "") if isinstance(who, dict) else ""
    except Exception:
        manager = ""
    if not manager:
        return "Не удалось определить менеджера по твоему Telegram — сделку не создаю."

    # Клиент: ищем по имени/телефону
    client_id, client_label, client_options = "", "", []
    if (client or "").strip():
        try:
            found = call("crm-clients.php", tg=str(tg_user_id),
                         params={"search": client.strip()})
            rows = found if isinstance(found, list) else (found or {}).get("clients", [])
            for c in rows[:5]:
                nm = (c.get("full_name")
                      or " ".join(p for p in (c.get("first_name"), c.get("last_name")) if p)
                      or c.get("company") or "?")
                contact = c.get("phone") or c.get("telegram") or c.get("email") or ""
                client_options.append((str(c.get("id")), f"{nm} ({contact})" if contact else nm))
        except Exception:
            pass
        if len(client_options) == 1:
            client_id, client_label = client_options[0]

    if not confirm:
        lines = [f"ПРЕДПРОСМОТР новой сделки (ещё НЕ создана):",
                 f"• Что: {title}",
                 f"• Сумма: {amount:,} ₽".replace(",", " ") if amount else "• Сумма: не указана",
                 f"• Менеджер: {manager}",
                 f"• Статус: {status}"]
        if client_id:
            lines.append(f"• Клиент: {client_label}")
        elif client_options:
            opts = "; ".join(f"{lbl} [id {cid}]" for cid, lbl in client_options)
            lines.append(f"• Клиент: найдено несколько — уточни: {opts}")
        elif client:
            lines.append(f"• Клиент: «{client}» в CRM не найден — могу создать без привязки, "
                         "НО ЛУЧШЕ привязать: спроси у менеджера номер телефона клиента — "
                         "по нему найду карточку или заведу новую (client_create) и привяжу")
        else:
            lines.append("• Клиент: не указан — спроси, на кого сделка (имя или телефон), "
                         "чтобы привязать к карточке")
        if notes:
            lines.append(f"• Заметка: {notes}")
        lines.append("Если всё верно — подтверди, и я создам (deal_create с confirm=true).")
        return "\n".join(lines)

    body = {"title": title.strip(), "status": status or "new",
            "assigned_to": manager, "created_by": manager, "currency": "RUB"}
    if amount:
        body["amount"] = int(amount)
    if client_id:
        body["client_id"] = client_id
    elif len(client_options) > 1:
        return "Найдено несколько клиентов — сначала уточни, который нужен (см. предпросмотр)."
    if notes:
        body["notes"] = notes
    if product_id:
        body["product_ids"] = str(product_id)
    try:
        res = call("crm-deals.php", tg=str(tg_user_id), method="POST", body=body)
    except Exception as e:
        return f"Не удалось создать сделку: {e}"
    # Дата продажи: проставляем closed_at/created_at на реальную дату сделки
    if sale_date:
        try:
            import re as _re
            if _re.fullmatch(r"\d{4}-\d{2}-\d{2}", sale_date.strip()):
                _did = res.get("id") if isinstance(res, dict) else None
                if _did:
                    call("crm-deals.php", tg=str(tg_user_id), method="PATCH",
                         body={"id": _did, "closed_at": sale_date.strip() + " 15:00:00",
                               "created_at": sale_date.strip() + " 15:00:00"})
        except Exception:
            pass
    deal_id = (res or {}).get("id", "?")
    client_txt = f", клиент {client_label}" if client_label else ""
    return f"Готово — сделка #{deal_id} создана: «{title}», менеджер {manager}{client_txt}."


@mcp.tool()
def make_selection_pdf(tg_user_id: str, items: list, title: str = "",
                       discount_percent: float = 0, client_name: str = "",
                       style_overrides: dict = None) -> str:
    """Собрать ПРЕЗЕНТАЦИЮ-PDF по фирменному шаблону StarGift и прислать файлом в Telegram.
    Используй, когда просят «подборку PDF/презентацию/файл для клиента».
    У каталожной позиции можно подменить главное фото: items[i]["photo_product"] =
    локальный путь/URL (напр. согласованная генерация exhibit_photo_frame) —
    цена и ссылка останутся из карточки.

    style_overrides — РАЗОВЫЕ правки вида для ЭТОЙ подборки (менеджер попросил
    «в этой сделай…»): {"price_bold": true|false, "show_link": bool,
    "show_final_slide": bool, "logo_position": "center"|"right",
    "text_size": "normal"|"large", "final_text": "свой текст финального слайда"}.
    НИКОГДА не отказывай «шаблон фиксированный» — правки вида ОБЯЗАТЕЛЬНЫ к
    исполнению. «Сделай так ВСЕГДА для меня» → инструмент selection_template_prefs.
    client_name — если менеджер назвал, ДЛЯ КОГО подборка (имя клиента), передай его:
    после отправки автоматически создастся напоминание-фоллоу-ап менеджеру через 3 дня
    («узнать решение по подборке») и заметка в карточке клиента. Не переспрашивай ради
    этого — передавай только если имя прозвучало.

    РЕГИСТР ТЕКСТА (правила дома StarGift, Вашик 22.07): title и любые подписи —
    люкс-тон (Sotheby's/Christie's/Rolex): спокойно, коротко, без шуток, капса,
    восклицаний, обращения на «ты». ⛔ ЗАПРЕТ темы подлинности на слайдах: не писать
    «подлинный/подлинность/сертификат подлинности». Сертификатор идёт мелкой
    спецификацией автоматически («Сертификат галереи · JSA/Beckett») — сам это не
    дублируй в тексте.

    Порядок работы: сначала подбери экспонаты через catalog_search (возьми оттуда id),
    затем для КАЖДОЙ позиции напиши blurb — 2–3 предложения, 150–300 знаков.

    СТРУКТУРА blurb — строго по приоритету (Вашик, 20.07.2026), абзацы через
    пустую строку, суммарно ДО ~600 знаков (иначе не влезет на слайд):
      1. Краткая сводка о подписанте — БУКВАЛЬНО ОДНО предложение (кто и чем велик).
      2. Расшифровка надписи/письма — ТОЛЬКО если она ДОСЛОВНО есть в карточке
         (catalog_get → fullDescription): «Текст дарственной надписи: «…»» /
         «Содержание письма: «…»» — цитатой, БЕЗ сокращений самой цитаты.
         НЕТ расшифровки в карточке → пункт просто пропусти: НЕ выдумывай её,
         НЕ пиши «расшифровка отсутствует» — сразу переходи к вещи.
      3. Совсем немного о самой вещи (дата, обстоятельства) — только если это
         реально добавляет ценности.
    Полную историю НЕ переносить простынёй — для этого внизу слайда есть ссылка
    «Больше информации об экспонате».

    НОВЫЕ ЗАПРЕТЫ (вердикты Вашика 22.07.2026, обязательные): слова «реликвия»,
    «под ключ», «частное собрание», «собирательный экспонат», «сопровождён спецификацией эксперта»;
    механические обороты («снимок фиксирует», «фотография демонстрирует»).
    Имена ВСЕГДА полные (Бен Хоган, не «Хоган»). Термины раскрывать («Тайгер-слэм» =
    четыре мэйджора подряд). Меньше тире: «— для кабинета» → «подходит для кабинета».
    Размеры НЕ в blurb (для них есть спецификация). Не выдумывать факты. Без списков
    и буллет-пойнтов. Слайды «о доме»/«сервис» клиенту не нужны — только экспонаты
    по теме. Не дублировать видимое на слайде («в наличии», «в раме» из названия). Blurb —
    ТОЛЬКО если добавляет контекст; самоочевидный лот («футболка с автографом»)
    идёт без blurb.

    СТИЛЬ blurb — строго ФАКТОЛОГИЧЕСКИЙ, как в фирменных подборках: кто такой,
    что и когда выиграл/сделал, чем именно знаменит предмет. ЗАПРЕЩЕНЫ лирика и
    продающие обороты («больше, чем футболист», «подарок, который не требует
    объяснений», «настоящая легенда», «уникальная возможность»).
    НЕ ПОВТОРЯЙ то, что уже написано на слайде: имя, тип предмета и число
    автографов видны в заголовке — blurb добавляет ТОЛЬКО новые факты
    (титулы, годы, чем знаменита вещь/момент). Плохо: «Автографы трёх участников
    Deep Purple на пластинке…» (дубль заголовка). Хорошо: «Классический состав,
    записавший In Rock и Machine Head. Концерт в Альберт-холле 1969 года — первое
    совместное выступление рок-группы с симфоническим оркестром.»
    Эталон: «Один из ключевых игроков легендарного состава Arsenal начала 2000-х
    годов. Именно Пирес был важной частью команды "Непобедимых", которая завершила
    чемпионат Англии без единого поражения.»
    Макет фиксирован — ты заполняешь только текст.

    items — список объектов:
      product_id (обяз. для позиций из каталога) — id карточки из catalog_search;
      blurb (обяз.) — твоя аргументация;
      cert (опц.) — бренд сертификации («Beckett», «JSA»), ТОЛЬКО если реально известен;
        пустая строка "" = общая формулировка о сертификации; не указывай для книг и
        экспонатов без автографа;
      photo_details (опц.) — свои URL фото-деталей из карточки, если дефолт не подходит.

    КАСТОМНАЯ позиция (экспоната нет на сайте, менеджер прислал фото): БЕЗ product_id,
    вместо этого укажи person (имя/название), headline (тип, напр. «Фото с автографом»),
    price (если назвали), blurb, и photo_product: "last" — возьмётся последнее
    присланное боту фото. Если фото не приложено — попроси прислать.

    Правки после отправки («поменяй описание у слайда N», «убери позицию») — просто
    вызови тул заново с теми же items и внесёнными изменениями: PDF пересобирается целиком.
    title — заголовок подборки, например «Подборка — футбол» (метаданные файла и подпись).
    discount_percent — персональная скидка клиента в процентах (напр. 15): на слайде
    останется обычная цена («Цена: …»), а главной строкой добавится «Цена с учётом
    вашей скидки: …» с пересчитанной суммой (процент на слайде НЕ печатается).
    Указывай ТОЛЬКО когда менеджер явно сказал про скидку этого клиента.
    """
    if not tg_user_id:
        return "Не указан tg_user_id."
    if not items or not isinstance(items, list):
        return "Передай items — список позиций {product_id, blurb, ...}."
    import selection_pdf_tool
    result = selection_pdf_tool.make_and_send(str(tg_user_id), items, title or "",
                                            discount_percent=discount_percent, style_overrides=style_overrides)
    # Follow-up: если подборка ДЛЯ конкретного клиента — напоминание менеджеру через 3 дня
    # + заметка в карточке. Ошибки фоллоу-апа не портят основной результат.
    if client_name and isinstance(result, str) and "не " not in result[:20].lower():
        from datetime import datetime, timedelta
        try:
            when = (datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d 11:00:00")
            create_reminder(tg_user_id, f"Фоллоу-ап: узнать решение по подборке — {client_name}",
                            when, client_name=client_name)
            result += f"\n(Поставил тебе фоллоу-ап по «{client_name}» через 3 дня, 11:00.)"
        except Exception:
            pass
        try:
            found = call("crm-clients.php", tg=str(tg_user_id),
                         params={"search": client_name, "limit": "1"})
            frows = found if isinstance(found, list) else (found.get("clients") or [])
            if frows:
                client_note_add(tg_user_id, str(frows[0]["id"]),
                                f"Отправлена PDF-подборка «{title or 'без названия'}» ({len(items)} поз.)")
        except Exception:
            pass
    return result


@mcp.tool()
def selection_template_prefs(tg_user_id: str, changes: dict = None, reset: bool = False) -> str:
    """ЛИЧНЫЕ настройки шаблона подборок менеджера — сохраняются НАВСЕГДА (пока не изменит).
    Используй, когда менеджер просит «сделай так ВСЕГДА / для меня по умолчанию»:
    жирная цена, без финального слайда, свой текст финального слайда, крупнее текст и т.п.
    changes: {"price_bold": bool, "show_link": bool, "show_final_slide": bool,
    "logo_position": "center"|"right", "text_size": "normal"|"large",
    "final_text": "личный текст финального слайда (можно с именем менеджера)"}.
    Без changes — показать текущие настройки. reset=true — вернуть заводской шаблон.
    Разовая правка одной подборки — это НЕ сюда, а style_overrides в make_selection_pdf."""
    if not tg_user_id:
        return "Не указан tg_user_id."
    import selection_prefs
    if reset:
        prefs = selection_prefs.reset_prefs(str(tg_user_id))
        return "Личные настройки сброшены к заводскому шаблону.\nДействующие: " + str(prefs)
    if changes:
        clean, errors = selection_prefs.validate(changes)
        if errors and not clean:
            return "Не понял настройки: " + "; ".join(errors)
        prefs = selection_prefs.set_prefs(str(tg_user_id), clean)
        note = (" (часть не принята: " + "; ".join(errors) + ")") if errors else ""
        return f"Запомнил — теперь твои подборки собираются так{note}:\n{prefs}"
    prefs = selection_prefs.get_prefs(str(tg_user_id))
    return "Твои текущие настройки шаблона:\n" + str(prefs)


@mcp.tool()
def person_texts_reference(tg_user_id: str, person: str) -> str:
    """ЛУЧШИЕ существующие описания персоны с наших же карточек — материал для нового
    описания. Вызывай ПЕРЕД catalog_add: собери из них богатое полное описание
    (биография + интересный факт), НЕ копируя дословно устаревшее (клуб/статус могли
    измениться — сверь и поправь; вечные факты: титулы, рекорды — бери смело)."""
    if not tg_user_id or not (person or "").strip():
        return "Нужно имя персоны."
    import base64
    import os as _os
    import subprocess as _sp
    b64 = base64.b64encode(person.strip().lower().encode()).decode()
    php = ("cd ~/stargift.ru/public_html/api && php8.2 -r '"
           "require \"db-config.php\"; $db=getDB();"
           f"$p=\"%\".base64_decode(\"{b64}\").\"%\";"
           "$st=$db->prepare(\"SELECT c.description d, COALESCE(t.text, c.full_description) tx "
           "FROM catalog c LEFT JOIN texts t ON t.product_id=c.product_id "
           "WHERE LOWER(c.person) LIKE ? AND c.is_published=1 "
           "ORDER BY LENGTH(COALESCE(t.text, c.full_description)) DESC LIMIT 2\");"
           "$st->execute([$p]);"
           "foreach($st->fetchAll(PDO::FETCH_ASSOC) as $r) "
           "echo \"=== \".$r[\"d\"].\" ===\\n\".mb_substr(strip_tags(str_replace(\"<br>\",\"\\n\",$r[\"tx\"])),0,2500).\"\\n\\n\";'")
    try:
        r = _sp.run(["ssh", "-i", "/Users/docbrown/.ssh/id_ed25519",
                     "stargift@stargift.beget.tech", php],
                    capture_output=True, text=True, timeout=60)
        out = "\n".join(l for l in r.stdout.split("\n") if "Welcome to" not in l).strip()
        if not out:
            return f"Готовых описаний «{person}» на сайте нет — собери описание сам (только проверенные факты)."
        return f"Существующие описания «{person}» (проверь актуальность клуба/статуса!):\n\n{out}"
    except Exception as e:
        return f"Не смог получить описания: {e}"


@mcp.tool()
def save_premium_feedback(tg_user_id: str, manager_name: str, feedback_text: str,
                          photo_paths: str = "", summary: str = "") -> str:
    """СОХРАНИТЬ фидбек менеджера по премиум-презентации для разработки.
    ОБЯЗАТЕЛЬНО вызывай при ЛЮБОМ отзыве/правках по премиальной презентации —
    ответ в чате записью НЕ считается. feedback_text — ДОСЛОВНЫЙ текст менеджера
    (не выжимка!), photo_paths — пути присланных фото через запятую + твоё
    описание в скобках, summary — твоё краткое резюме сути."""
    if not feedback_text.strip():
        return "Пустой фидбек."
    import time as _t
    stamp = _t.strftime("%Y-%m-%d %H:%M")
    entry = (f"\n## {stamp} — {manager_name or tg_user_id} — премиум-презентация\n"
             f"Дословно: «{feedback_text.strip()}»\n"
             f"Фото: {photo_paths.strip() or 'нет'}\n"
             f"Суть: {summary.strip() or '—'}\n")
    with open("/Users/docbrown/hermes-doc/feedback/premium-feedback.md", "a") as f:
        f.write(entry)
    return "Фидбек сохранён для разработки. Поблагодари менеджера."


@mcp.tool()
def weekend_reminders(tg_user_id: str, mute: bool = True) -> str:
    """Пауза напоминаний по ВЫХОДНЫМ для менеджера (сб+вс): mute=True — не слать
    в выходные планёрку и пуши-напоминания (они не теряются, приходят в понедельник);
    mute=False — вернуть как было. Вызывай, когда менеджер просит «не присылай
    напоминания по субботам/воскресеньям» или «верни напоминания в выходные».
    Действует для написавшего менеджера (tg_user_id из сессии)."""
    if not tg_user_id:
        return "Нужен tg_user_id."
    import json as _json, re as _re, subprocess as _sp
    tg = str(tg_user_id).strip()
    if not _re.fullmatch(r"\d{5,15}", tg):
        return "Некорректный id."
    # 1) локальные префы (планёрка)
    pf = "/Users/docbrown/hermes-doc/notify_prefs.json"
    try:
        prefs = _json.load(open(pf))
    except Exception:
        prefs = {}
    lst = set(prefs.get("weekend_mute", []))
    (lst.add(tg) if mute else lst.discard(tg))
    prefs["weekend_mute"] = sorted(lst)
    _json.dump(prefs, open(pf, "w"), ensure_ascii=False, indent=1)
    # 2) серверные пуши: tg_links → crm_user_id → tg_notif_prefs.weekend_mute
    val = 1 if mute else 0
    php = ("cd ~/stargift.ru/public_html/api && php8.2 -r '"
           "require \"db-config.php\"; $db=getDB();"
           f"$st=$db->prepare(\"SELECT crm_user_id FROM tg_links WHERE tg_user_id=? LIMIT 1\"); $st->execute([{tg}]);"
           "$uid=$st->fetchColumn(); if(!$uid){echo \"NOLINK\"; exit;}"
           "try{$db->exec(\"ALTER TABLE tg_notif_prefs ADD COLUMN weekend_mute TINYINT(1) NOT NULL DEFAULT 0\");}catch(Exception $e){}"
           "$db->prepare(\"INSERT INTO tg_notif_prefs (crm_user_id, weekend_mute) VALUES (?, ?) "
           f"ON DUPLICATE KEY UPDATE weekend_mute={val}\")->execute([$uid, {val}]);"
           "echo \"OK\";'")
    r = _sp.run(["ssh", "-i", "/Users/docbrown/.ssh/id_ed25519", "stargift@stargift.beget.tech", php],
                capture_output=True, text=True, timeout=60)
    out = (r.stdout or "").strip().splitlines()[-1] if r.stdout.strip() else ""
    if out == "NOLINK":
        return ("Локально сделал, но менеджер не привязан в CRM (tg_links) — пуши браузера "
                "настроить не смог. Планёрки в выходные уже " + ("отключены." if mute else "включены."))
    if out != "OK":
        return f"Планёрки настроил, но серверная настройка пушей не прошла: {r.stderr[:120] or out}"
    return ("Готово: по субботам и воскресеньям напоминания и планёрки приходить не будут — "
            "всё накопившееся придёт в понедельник." if mute else
            "Готово: напоминания в выходные снова включены.")


@mcp.tool()
def exhibit_photo_frame(tg_user_id: str, photo_path: str, person_latin: str = "",
                        style_notes: str = "", frame_style: str = "",
                        poster_path: str = "", plate_title: str = "",
                        plate_sub: str = "") -> str:
    """Оформить РЕАЛЬНОЕ фото экспоната «в раму» (наш фирменный вид: чёрная рама,
    синее паспарту с золотой окантовкой, золотой шильд с именем) и ПРИСЛАТЬ результат
    в чат на согласование. Сам предмет НЕ перерисовывается — только окружение.
    photo_path — путь к фото из сообщения («The file is saved at: …»), локальный путь
    к фото карточки, https-URL фото (например img лота из get_last_presentation —
    «оформи фото Бекхэма из подборки в раму»), ЛИБО строка "last" — правка ПОСЛЕДНЕЙ
    генерации (используй, когда менеджер отвечает на присланное ботом фото оформления:
    «на чёрном фоне», «шильд больше» → photo_path="last", style_notes=правка).
    plate_title/plate_sub — для экспонатов БЕЗ АВТОГРАФА (гравюры, старинные книги,
    документы без подписи): на шильд идёт НАЗВАНИЕ в кавычках (plate_title) и ниже
    мелко год + издание (plate_sub, напр. «1601, Георг Браун и Франс Хогенберг»;
    для книг — год и издательство). Правило Вашика 31.07: пустых шильдов не бывает —
    есть автограф → person_latin, нет автографа → plate_title + plate_sub.
    person_latin — имя ЛАТИНИЦЕЙ для шильда («LIONEL MESSI»). ПЕРЕДАВАЙ ВСЕГДА
    (стандарт приёмки: шильд не бывает пустым): персона известна из карточки/контекста —
    транслитерируй общепринятым написанием; команда — название команды латиницей.
    Пусто ТОЛЬКО если менеджер явно попросил без имени или персона неизвестна
    (фото из чата без карточки — уточни одной строкой). Гравировка локальная,
    генератору имя не передаётся — блокировок не бывает. frame_style — КЛЮЧ рамы из УТВЕРЖДЁННОГО списка (наши реальные рамы, выдумывать
    НЕЛЬЗЯ): "brand" (фирменная чёрная+синее паспарту, дефолт для фото), "photo-brown-velvet"
    (коричневая+бордовый бархат, музыка), "retro-gold-baroque" (золотая резная+бордовый
    бархат — дорогое ретро), "retro-silver-pearl" (серебристо-жемчужная+синий бархат —
    ретро-актрисы), "glass-float" (между стёкол БЕЗ паспарту — постеры/афиши/документы),
    "tshirt-black-flat" (чёрная плоская — стандарт футболок), "packshot-white"/"packshot-black"
    (ПРЕДМЕТ на чистом белом/чёрном студийном фоне БЕЗ рамы — для бутс/мячей/предметов
    на некрасивом фоне; шильда нет, person_latin не нужен),
    "historic-brown-gold" (коричневая с золотым кантом — исторические),
    "historic-silver-blue" (серебристая+синий бархат), "vinyl-gold-velvet"
    (золотая фактурная+бархат — пластинки), "cinema-composite" (КИНО-КОМПОЗИТ:
    предмет + постер фильма в одной раме с бархатом — для дорогих кино-лотов;
    ОБЯЗАТЕЛЕН poster_path — второе фото/URL с постером фильма).
    Выбирай по типу экспоната автоматически. «Несколько вариантов рам» = несколько
    вызовов с РАЗНЫМИ ключами из списка (не фантазийные описания!).
    style_notes — только детали НЕ про раму («тёмный фон», «без шильда») по-русски.
    После согласования менеджером: фото карточки — через catalog_photo_update;
    фото в ПОДБОРКЕ — пересобери make_presentation_premium, подставив путь результата
    в photo_url лота (локальные пути поддерживаются)."""
    if not tg_user_id or not photo_path:
        return "Нужны tg_user_id и photo_path."
    import framing_tool
    import os as _os
    if photo_path.startswith("http://") or photo_path.startswith("https://"):
        # фото лота из подборки/каталога — скачиваем во временный файл
        import tempfile as _tf
        import urllib.request as _rq
        url = photo_path.replace("https://media.stargift.ru/", "https://stargift.ru/media/")
        url = url.replace("s-l500", "s-l1600")
        try:
            req = _rq.Request(url, headers={"User-Agent": "Mozilla/5.0 (StargiftBot)"})
            data = _rq.urlopen(req, timeout=60).read()
            ext = ".png" if url.lower().split("?")[0].endswith(".png") else ".jpg"
            fd, tmp = _tf.mkstemp(suffix=ext, prefix="frame_src_")
            with _os.fdopen(fd, "wb") as fh:
                fh.write(data)
            photo_path = tmp
        except Exception as e:
            return f"Не смог скачать фото по ссылке: {e}"
    if not _os.path.exists(photo_path):
        return f"Файл не найден: {photo_path}"
    poster_local = ""
    if poster_path:
        if poster_path.startswith("http://") or poster_path.startswith("https://"):
            import tempfile as _tf2
            import urllib.request as _rq2
            u2 = poster_path.replace("https://media.stargift.ru/", "https://stargift.ru/media/")
            try:
                rq2 = _rq2.Request(u2, headers={"User-Agent": "Mozilla/5.0 (StargiftBot)"})
                dat2 = _rq2.urlopen(rq2, timeout=60).read()
                fd2, poster_local = _tf2.mkstemp(suffix=".jpg", prefix="frame_poster_")
                with _os.fdopen(fd2, "wb") as fh2:
                    fh2.write(dat2)
            except Exception as e:
                return f"Не смог скачать постер по ссылке: {e}"
        elif _os.path.exists(poster_path):
            poster_local = poster_path
        else:
            return f"Файл постера не найден: {poster_path}"
    try:
        if frame_style == "cinema-composite":
            if not poster_local:
                return "Для кино-композита нужен poster_path (фото/URL постера фильма)."
            res = framing_tool.frame_cinema_composite(photo_path, poster_local, person_latin, style_notes)
            if res.get("error"):
                return res["error"]
        else:
            res = framing_tool.frame_exhibit_photo(photo_path, person_latin, style_notes,
                                                   frame_style, plate_title, plate_sub)
    except Exception as e:
        return f"Генерация не удалась: {e}. Можно попробовать ещё раз или позвать Вашика."
    note = "" if res["shield_engraved"] or not person_latin else " (шильд-табличку найти не смог — имя не нанесено)"
    sent = framing_tool.send_photo(str(tg_user_id), res["path"],
                                   "Вариант оформления — подтверди или скажи, что поправить.")
    if sent:
        return f"Фото оформления отправлено в чат{note}. Файл: {res['path']}. После «да» менеджера — ставь в карточку через catalog_photo_update."
    return f"Сгенерировал, но не смог отправить в чат{note}. Файл: {res['path']}."


@mcp.tool()
def catalog_photo_update(tg_user_id: str, product_query: str, photo_path: str = "",
                         mode: str = "replace_main", confirm: bool = False) -> str:
    """Обновить ФОТО карточки товара. mode: replace_main (заменить главное),
    add (добавить вторым), remove_extra (убрать все фото кроме главного),
    replace_all (оставить только новое фото). photo_path — локальный путь к новому фото
    (из чата или из exhibit_photo_frame), ЛИБО "last" — ПОСЛЕДНЕЕ сгенерированное
    оформление («поставь то, что ты сгенерировал»). Для remove_extra не нужен.
    БЕЗ confirm — предпросмотр (что изменится); менеджер сказал «да» → confirm=true.
    product_query — название, ссылка на карточку или id (sg_… / цифры)."""
    if not tg_user_id or not product_query:
        return "Нужны tg_user_id и product_query."
    if mode not in ("replace_main", "add", "remove_extra", "replace_all"):
        return "mode: replace_main | add | remove_extra | replace_all"
    import os as _os, re as _re, subprocess as _sp, time as _time
    if photo_path == "last":
        import framing_tool
        photo_path = framing_tool.latest_render() or ""
        if not photo_path:
            return "Сгенерированных оформлений не нашёл — сначала exhibit_photo_frame."
    q = str(product_query).strip()
    m = _re.search(r"(sg_\d+_[a-f0-9]+)", q) or _re.search(r"/product/(\d+)", q)
    pid = m.group(1) if m else (q if (q.startswith("sg_") or q.isdigit()) else "")
    if not pid:
        found = _fetch_products(q, 3, 0)
        if len(found) == 1:
            pid = str(found[0].get("id") or "")
        elif len(found) > 1:
            names = "; ".join(str(p.get("title") or p.get("person") or p.get("id")) for p in found[:3])
            return f"Нашёл несколько карточек: {names}. Уточни, какая именно (лучше ссылкой)."
        else:
            return "Карточка не найдена — пришли ссылку на неё."
    ssh = ["ssh", "-i", "/Users/docbrown/.ssh/id_ed25519", "stargift@stargift.beget.tech"]
    php_get = ("cd ~/stargift.ru/public_html/api && php8.2 -r '"
               "require \"db-config.php\"; $db=getDB(); $st=$db->prepare(\"SELECT photo_url FROM catalog WHERE product_id=?\");"
               f"$st->execute([\"{pid}\"]); echo $st->fetchColumn();'")
    cur = _sp.run(ssh + [php_get], capture_output=True, text=True, timeout=60).stdout.strip().split("\n")[-1]
    photos = [p for p in cur.split() if p.startswith("http")]
    if mode in ("replace_main", "add", "replace_all"):
        if not photo_path or not _os.path.exists(photo_path):
            return f"Файл нового фото не найден: {photo_path}"
        fname = f"exh_{pid[-8:]}_{int(_time.time())}.jpg"
        new_url = f"https://stargift.ru/media/{fname}"
        if mode == "replace_main":
            new_photos = [new_url] + photos[1:]
        elif mode == "add":
            new_photos = photos + [new_url]
        else:
            new_photos = [new_url]
    else:
        if len(photos) <= 1:
            return "У карточки и так одно фото — убирать нечего."
        new_photos = photos[:1]
    if not confirm:
        return (f"ПРЕДПРОСМОТР для #{pid}: фото станет {len(new_photos)} шт. "
                f"(было {len(photos)}), режим {mode}. Подтверди — применю (confirm=true).")
    if mode != "remove_extra":
        up = _sp.run(["scp", "-i", "/Users/docbrown/.ssh/id_ed25519", photo_path,
                      f"stargift@stargift.beget.tech:~/stargift.ru/public_html/media/{fname}"],
                     capture_output=True, text=True, timeout=120)
        if up.returncode != 0:
            return f"Не смог загрузить фото на сервер: {up.stderr[:200]}"
    joined = " ".join(new_photos)
    php_set = ("cd ~/stargift.ru/public_html/api && php8.2 -r '"
               "require \"db-config.php\"; $db=getDB(); $st=$db->prepare(\"UPDATE catalog SET photo_url=? WHERE product_id=?\");"
               f"$st->execute([\"{joined}\", \"{pid}\"]); echo \"ok\";'")
    r = _sp.run(ssh + [php_set], capture_output=True, text=True, timeout=60)
    if "ok" not in r.stdout:
        return f"Ошибка обновления карточки: {(r.stderr or r.stdout)[:200]}"
    return f"Готово: у карточки #{pid} теперь {len(new_photos)} фото. Проверить: https://stargift.ru/product/{pid}/"


@mcp.tool()
def vedenie_report(tg_user_id: str, month: str = "", manager: str = "") -> str:
    """Сводка по ВЕДЕНИЯМ — первичному сопровождению новых клиентов (2000 ₽ менеджеру
    за каждое). Ведение = первая сделка клиента, который ещё не был в чьей-либо базе;
    помечается автоматически при создании сделки. После ведения клиент закрепляется
    за менеджером, дальнейшие продажи идут в её план как обычные.
    month: YYYY-MM (пусто = текущий). manager: имя — только её ведения (менеджеру
    показывай ТОЛЬКО её собственные; всю сводку — только Вашику/директору)."""
    if not tg_user_id:
        return "Не указан tg_user_id."
    params = {}
    if month:
        params["month"] = month
    if manager:
        params["manager"] = manager
    data = call("vedenie-report.php", tg=str(tg_user_id), params=params)
    if not isinstance(data, dict) or data.get("error"):
        return "Не смог получить сводку по ведениям."
    lines = [f"Ведения за {data.get('month')}: итого выплат {data.get('total_fee', 0)} ₽"]
    for m in data.get("managers", []):
        lines.append(f"\n{m['manager']}: {m['count']} шт. × 2000 = {m['fee_total']} ₽")
        for d in m.get("deals", [])[:10]:
            lines.append(f"  • {d['title']} ({d['created_at']})")
    if len(data.get("managers", [])) == 0:
        lines.append("Ведений в этом месяце пока нет.")
    return "\n".join(lines)


@mcp.tool()
def selection_template_from_example(tg_user_id: str, file_path: str, apply: bool = False) -> str:
    """Обучение шаблона подборок на ЭТАЛОНЕ менеджера. Используй, когда менеджер
    ПРИСЛАЛ ФАЙЛ .key или .pptx (путь будет в сообщении: «The file is saved at: …»)
    и просит «хочу такой шаблон / сделай как у меня в примере».
    Двухшаговый сценарий: (1) вызови с apply=false — получишь, какие настройки
    распознаны; ПЕРЕСКАЖИ их менеджеру по-русски и спроси подтверждение;
    (2) после «да» вызови ПОВТОРНО с apply=true — настройки сохранятся навсегда.
    Отличия, которые опции не покрывают, честно назови и пообещай передать
    на доработку шаблона (Вашик их увидит)."""
    if not tg_user_id or not file_path:
        return "Нужны tg_user_id и путь к файлу."
    import template_from_example
    res = template_from_example.analyze(file_path)
    if res.get("error"):
        return "Не смог разобрать эталон: " + res["error"]
    proposed, unmapped = res.get("proposed") or {}, res.get("unmapped") or []
    lines = []
    if proposed:
        lines.append("Распознанные настройки: " + str(proposed))
    else:
        lines.append("Отличий, которые я умею запоминать, в эталоне не нашёл.")
    if unmapped:
        lines.append("Вне моих опций (передам на доработку): " + "; ".join(unmapped))
    if apply and proposed:
        import selection_prefs
        prefs = selection_prefs.set_prefs(str(tg_user_id), proposed)
        lines.append("ПРИМЕНЕНО НАВСЕГДА. Действующие настройки: " + str(prefs))
    elif proposed:
        lines.append("Пока НЕ применено — перескажи менеджеру и после «да» вызови с apply=true.")
    if unmapped:
        try:
            import json as _json, time as _time
            with open("/Users/docbrown/hermes-doc/template_requests.log", "a") as f:
                f.write(_json.dumps({"ts": _time.strftime("%Y-%m-%d %H:%M"),
                                     "tg": str(tg_user_id), "file": file_path,
                                     "unmapped": unmapped}, ensure_ascii=False) + "\n")
        except Exception:
            pass
    return "\n".join(lines)


@mcp.tool()
def catalog_add(tg_user_id: str, person: str = "", name: str = "", price: int = 0,
                exhibit_type: str = "", description: str = "", short_description: str = "",
                gallery: str = "", categories: str = "", with_autograph: bool = True,
                confirm: bool = False, parent_product_id: str = "",
                image_path: str = "") -> str:
    """Создать НОВЫЙ экспонат в каталоге сайта. БЕЗ confirm=true — ПРЕДПРОСМОТР (не создаёт);
    с confirm=true — создаёт.

    Фото: обычно берётся последнее присланное в чат; если фото пришло ФАЙЛОМ-документом —
    передай image_path из сообщения («The file is saved at: …»); image_path="last" —
    последняя генерация оформления (после exhibit_photo_frame).

    ВАРИАНТ существующей карточки («мяч на подставке», «мрамор/дерево», «в раме»):
    передай parent_product_id = product_id ОСНОВНОЙ карточки (найди через catalog_search
    или возьми из ссылки), person = «{имя как у родителя} - {Вариант}» (пример:
    «Хабиб Нурмагомедов - Дерево»), своё фото и цену. Вариант появится в селекторе
    оформления на карточке родителя.

    Обязательно: person (чей автограф/кто), name (название), exhibit_type (рама/подставка/как есть),
    и ФОТО — берётся из последнего присланного пользователем изображения (попроси прислать, если нет).
    Фото обрабатывается автоматически: главное → формат 4:3 для сайта; вторым добавляется крупный
    план автографа (отключить — with_autograph=false, если подписи на фото нет).
    Необязательно:
      • price (₽);
      • description — ПОДРОБНОЕ описание экспоната (идёт в полное описание карточки);
      • short_description — короткий подзаголовок/тип («Футболка с автографом») — НЕ дублируй сюда
        подробный текст; если не задан, подставится название;
      • gallery — ТОЛЬКО если пользователь явно назвал (Времена года/Гименей/Dream House/
        Барвиха/Stargift Prime). НЕ ЗНАЕШЬ — оставь ПУСТЫМ, не выдумывай заглушки
        («неизвестно», «нет данных» и т.п. запрещены — они попадут на сайт буквально);
      • categories (через ;).
    Публикация по роли: менеджер+ — сразу в каталог; junior/staff — на согласование.
    tg_user_id — Telegram id из контекста.
    """
    if not tg_user_id:
        return "Не указан tg_user_id."
    import catalog_create
    import interior_fit

    missing = catalog_create.missing_fields(person, name, exhibit_type)
    if missing:
        return "Чтобы добавить экспонат, не хватает: " + ", ".join(missing) + ". Уточни у пользователя."

    # Фото: явный путь (из «The file is saved at: …» или "last" = последняя генерация
    # оформления) имеет приоритет; иначе — последнее присланное в чат фото из кэша.
    import os as _os
    if image_path == "last":
        import framing_tool
        image_path = framing_tool.latest_render() or ""
    if image_path and not _os.path.exists(image_path):
        image_path = ""
    if not image_path:
        image_path = interior_fit.latest_incoming_image(set())
    if not image_path:
        return "Нужно фото экспоната — попроси прислать фото одним сообщением (файлом тоже подходит)."

    et = catalog_create.normalize_exhibit_type(exhibit_type)
    type_ru = {"frame": "в раме", "stand": "на подставке", "as_is": "как есть"}.get(et, et)
    if not confirm:
        return (
            "ПРЕДПРОСМОТР нового экспоната:\n"
            f"• Название: {name}\n• Автор/кто: {person}\n• Тип: {type_ru}\n"
            f"• Цена: {int(price or 0)} ₽\n• Галерея: {gallery or '—'}\n"
            f"• Категории: {categories or '—'}\n• Фото: есть\n"
            "Если всё верно — подтверди, и я создам (catalog_add с confirm=true)."
        )

    res = catalog_create.create_exhibit(str(tg_user_id), person, name, price, et,
                                        image_path, description=short_description,
                                        gallery=gallery, categories=categories,
                                        full_description=description,
                                        with_autograph=with_autograph,
                                        parent_product_id=parent_product_id)
    if not isinstance(res, dict) or res.get("error"):
        return f"Не удалось создать экспонат: {res.get('error') if isinstance(res, dict) else res}"
    status = "опубликован в каталоге" if res.get("moderation_status") == "published" else "отправлен на согласование старшему"
    pid = res.get("product_id")
    return f"Готово — экспонат «{name}» {status}." + (f" id {pid}" if pid else "")


# ── Менеджерский паритет: клиенты, заявки, напоминания, шаблоны ──────────────

def _manager_name(tg: str) -> str:
    """Имя сотрудника по его Telegram id (bot-resolve-user)."""
    try:
        who = call("bot-resolve-user.php", tg=str(tg), params={"tg": str(tg)})
        return (who or {}).get("name", "") if isinstance(who, dict) else ""
    except Exception:
        return ""


@mcp.tool()
def tops_exhibits_watch(tg_user_id: str, view: str = "sheet", person: str = "") -> str:
    """ТОП-экспонаты: источник правды — Google-таблица «НАЛИЧИЕ ТОПЫ» + сверка с сайтом.

    view="sheet" (по умолчанию) — СЖАТАЯ сводка из ТАБЛИЦЫ + сверка с сайтом:
    сначала проблемы (нет карточек на сайте / остался 1 шт / мало), потом наличие
    по персонам компактной строкой. Используй на «что по топам», «проверь топы».
    person="Месси" — ДЕТАЛИ по одной персоне: конкретные предметы по галереям.
    Используй, когда спрашивают про конкретного («что по топам Месси»).
    view="full" — развёрнутый список всех персон по галереям.
    view="crm" — запасной режим по складу CRM: дозаказать/оформить/перевезти;
    view="galleries" — склад CRM по галереям.
    tg_user_id — Telegram id из контекста."""
    if not tg_user_id:
        return "Не указан tg_user_id."

    if str(view).lower() in ("sheet", "full", "sheet_full", "таблица", "table", "") or person:
        try:
            import tops_sheet
            persons = tops_sheet.parse_tops(tops_sheet.download_sheet())
            site = tops_sheet.fetch_site_catalog()
            res = tops_sheet.compare_with_site(persons, site)
        except Exception as e:
            return (f"Не смог прочитать таблицу «НАЛИЧИЕ ТОПЫ» ({e}). "
                    "Могу показать по складу CRM — скажи «топы по складу».")
        total = sum(p["total"] for p in res)

        # ── Деталка по одной персоне ──
        if person:
            q = tops_sheet._fold(person)
            hits = [p for p in res if q in tops_sheet._fold(p["person"])]
            if not hits:
                return f"«{person}» в таблице топов не нашёл. Есть: " + ", ".join(x["person"] for x in res[:15]) + "…"
            out = []
            for p in hits:
                out.append(f"📌 {p['person']} — {p['total']} шт | на сайте {p['site_cards']} живых карточек")
                for g, items in p["galleries"].items():
                    if items:
                        out.append(f"  {g}:")
                        out += [f"   • {label}" + (f" — {n} шт" if n != 1 else "") for label, n in items]
            return "\n".join(out)

        # ── Развёрнутый список ──
        if str(view).lower() in ("full", "sheet_full"):
            lines = [f"ТОПЫ (таблица «НАЛИЧИЕ ТОПЫ»): {len(res)} персон, {total} шт."]
            for p in res:
                gal = ", ".join(f"{g} {sum(n for _, n in items)}"
                                for g, items in p["galleries"].items() if items)
                site_mark = f"сайт {p['site_cards']}" if p["site_cards"] else "сайт: НЕТ ⚠️"
                lines.append(f"• {p['person']} — {p['total']} шт ({gal}) | {site_mark}")
            return "\n".join(lines)

        # ── Сжатая сводка (дефолт) ──
        lines = [f"ТОПЫ: {len(res)} персон, {total} шт (таблица «НАЛИЧИЕ ТОПЫ» + сверка с сайтом)"]
        missing = [tops_sheet.pretty_person(p["person"]) for p in res if p["site_cards"] == 0]
        low1 = [tops_sheet.pretty_person(p["person"]) for p in res if p["total"] == 1]
        low23 = [f"{tops_sheet.pretty_person(p['person'])} ({p['total']})" for p in res if 2 <= p["total"] <= 3]
        if missing:
            lines.append("⚠️ Нет живых карточек на сайте:")
            lines += [f"• {x}" for x in missing]
        if low1:
            lines.append("🔴 Остался 1 шт:")
            lines += [f"• {x}" for x in low1]
        if low23:
            lines.append("🟡 Мало (2–3 шт):")
            lines += [f"• {x}" for x in low23]
        if not (missing or low1 or low23):
            lines.append("✅ Проблем нет: наличие и карточки в порядке.")
        by_cat = {}
        for p in sorted(res, key=lambda x: -x["total"]):
            by_cat.setdefault(tops_sheet.category(p["person"]), []).append(p)
        lines.append("\n📦 Наличие (шт):")
        for cat in ("Спорт", "Кино", "Музыка", "Другое"):
            if cat in by_cat:
                lines.append(f"\n{cat}:")
                lines += [f"• {tops_sheet.pretty_person(p['person'])} — {p['total']}"
                          for p in by_cat[cat]]
        lines.append("\nДетали по персоне: спроси «топы Месси». Полный разбор — «топы подробно».")
        return "\n".join(lines)
    try:
        data = call("exhibits.php", tg=str(tg_user_id))
    except Exception as e:
        return f"Ошибка доступа к складу экспонатов: {e}"
    rows = data if isinstance(data, list) else (data or {}).get("exhibits", [])
    tops = [r for r in rows
            if (r.get("is_top") or "ТОП" in str(r.get("category") or "").upper())
            and not r.get("crm_hidden")]
    if not tops:
        return "ТОП-экспонаты на складе не найдены."

    # Поле person часто содержит хвост про предмет («Аль Пачино Кубок с автографом»)
    # — нормализуем до имени, чтобы считать НАЛИЧИЕ ПО ПЕРСОНЕ, а не по вариациям.
    _ITEM_WORDS = ("фото", "автограф", "кубок", "брошюра", "статуэтк", "мяч", "бутса",
                   "кроссовк", "рисун", "книга", "письмо", "банкнот", "карточк",
                   "постер", "сценари", "перчатк", "плакат", "обложк", "журнал",
                   "диплом", "димлом", "футболк", "бейсболк", "клюшк", "шлем", "ракетк",
                   "гитар", "белая", "белый", "боксерск", "боксёрск", "коллекционн",
                   "подписан", "-", "—", "(")

    def _person_key(raw: str) -> str:
        words = (raw or "?").split()
        out = []
        for w in words:
            lw = w.lower().strip(".,«»\"'()")
            if any(lw.startswith(iw) for iw in _ITEM_WORDS):
                break
            out.append(w)
        return " ".join(out) or (raw or "?")

    STORAGE = ("сейф", "вне галерей", "без галереи", "склад")
    groups: dict = {}
    for e in tops:
        raw_person = str(e.get("person") or "?")
        if "продан" in raw_person.lower():
            continue  # помечен проданным — в наличии не считаем
        g = groups.setdefault(_person_key(raw_person), {
            "total": 0, "framed": 0, "unframed": 0, "in_framing": 0, "galleries": {}})
        g["total"] += 1
        fs = e.get("framing_status")
        g["framed" if fs == "framed" else "in_framing" if fs == "in_framing" else "unframed"] += 1
        gal = (e.get("gallery") or "без галереи").strip() or "без галереи"
        g["galleries"][gal] = g["galleries"].get(gal, 0) + 1

    if str(view).lower().startswith("galler") or "галере" in str(view).lower():
        by_gal: dict = {}
        for person, g in groups.items():
            for gal, n in g["galleries"].items():
                by_gal.setdefault(gal, []).append((person, n, g["unframed"]))
        def _gal_rank(name: str) -> tuple:
            return (any(st in name.lower() for st in STORAGE), name)
        lines = [f"ТОП-экспонаты по галереям: {len(groups)} персон, {len(tops)} шт."]
        for gal in sorted(by_gal, key=_gal_rank):
            items = sorted(by_gal[gal])
            total = sum(n for _, n, _ in items)
            lines.append(f"\n📍 {gal} — {total} шт:")
            lines += [f"• {person} — {n} шт" for person, n, _ in items]
        return "\n".join(lines)

    reorder, frame, move, expose = [], [], [], []
    for person, g in sorted(groups.items()):
        gal_str = ", ".join(f"{k}: {v}" for k, v in sorted(g["galleries"].items()))
        showcase = {k: v for k, v in g["galleries"].items()
                    if not any(st in k.lower() for st in STORAGE)}
        if g["total"] <= 1:
            reorder.append(f"• {person} — осталось {g['total']} шт ({gal_str})")
        if g["unframed"] > 0:
            frame.append(f"• {person} — не оформлено {g['unframed']} из {g['total']}"
                         + (f" (в багетной: {g['in_framing']})" if g["in_framing"] else ""))
        if g["total"] >= 2 and len(showcase) == 1:
            only = next(iter(showcase))
            move.append(f"• {person} — в витринах только «{only}» ({gal_str}) — рассмотреть перевозку")
        if g["total"] >= 1 and not showcase:
            expose.append(f"• {person} — ни одного в витринах галерей ({gal_str})")

    lines = [f"ТОП-экспонаты склада: {len(groups)} персон, {len(tops)} шт. "
             f"(наличие считается по персоне; уникальные вещи в «дозаказать» = кандидаты на пополнение той же персоной)"]
    if reorder:
        lines.append("\n🔴 МАЛО В НАЛИЧИИ (≤1 по персоне) — дозаказать/пополнить:\n" + "\n".join(reorder))
    if frame:
        lines.append("\n🟡 ОФОРМИТЬ (не в раме):\n" + "\n".join(frame))
    if expose:
        lines.append("\n🟣 ВЫСТАВИТЬ (всё в сейфе/вне галерей):\n" + "\n".join(expose))
    if move:
        lines.append("\n🔵 ПЕРЕВЕЗТИ (в витринах только одна галерея):\n" + "\n".join(move))
    if not (reorder or frame or move or expose):
        lines.append("Всё в порядке: наличие, оформление и распределение без замечаний.")
    return "\n".join(lines)


def _rub(v) -> str:
    try:
        return f"{int(float(v)):,} ₽".replace(",", " ")
    except (TypeError, ValueError):
        return str(v or 0)


def _days_since(date_str) -> int:
    """Сколько дней прошло с даты (ISO/‘YYYY-MM-DD’). 9999 если неизвестно."""
    if not date_str:
        return 9999
    from datetime import datetime
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
        try:
            d = datetime.strptime(str(date_str)[:19], fmt)
            return max(0, (datetime.now() - d).days)
        except ValueError:
            continue
    return 9999


def _days_until_event(date_str) -> int:
    """Дней до ближайшего годового повтора даты (ДР/годовщина). 9999 если не распарсить."""
    if not date_str:
        return 9999
    from datetime import date, datetime
    d = None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m", "%m-%d"):
        try:
            d = datetime.strptime(str(date_str)[:10], fmt)
            break
        except ValueError:
            continue
    if d is None:
        return 9999
    today = date.today()
    try:
        nxt = date(today.year, d.month, d.day)
    except ValueError:
        return 9999  # 29 февраля и т.п. — не усложняем
    if nxt < today:
        try:
            nxt = date(today.year + 1, d.month, d.day)
        except ValueError:
            return 9999
    return (nxt - today).days


def _client_label(c: dict) -> str:
    name = (c.get("name") or c.get("full_name") or "").strip()
    if name:
        return name
    phone = c.get("phone") or ""
    if phone:
        return f"тел. {phone}"
    cid = str(c.get("id") or "")
    return f"клиент {cid[:8]}" if cid else "клиент без имени"


def _scan_clients(tg_user_id: str, max_scan: int = 1500) -> list:
    """Постранично тянем клиентов (сервер сам ограничивает менеджера его базой)."""
    out, offset, page = [], 0, 500
    while len(out) < max_scan:
        c = call("crm-clients.php", tg=str(tg_user_id),
                 params={"limit": str(page), "offset": str(offset)})
        rows = c if isinstance(c, list) else (c.get("clients") or c.get("data") or [])
        if not rows:
            break
        out.extend(rows)
        if len(rows) < page:
            break
        offset += page
    return out[:max_scan]


@mcp.tool()
def clients_to_target(tg_user_id: str, mode: str = "reactivate",
                      query: str = "", days: int = 14, limit: int = 15) -> str:
    """Кому стоит направить подборку — аналитика клиентской базы (кандидаты на проактивный контакт).

    mode:
      reactivate — ценные, но затихшие клиенты (RFM-сегменты «под угрозой» + «спящие»).
                   Кому пора напомнить о себе новой подборкой. Это режим по умолчанию.
      vip        — крупные разовые покупатели (RFM «крупные разовые») — кандидаты на персональный VIP-оффер.
      occasion   — у кого скоро повод (ДР/годовщина/именины в ближайшие `days` дней) — подборка к дате.
      theme      — кому близка тема/персона из `query` (по интересам/тегам/досье). Напр. query="Месси", "бокс".

    Менеджеру сервер отдаёт только ЕГО клиентов; директору/owner/admin — по всей базе.
    days — горизонт для mode=occasion (по умолчанию 14). query — тема для mode=theme.
    limit — сколько кандидатов показать (по умолчанию 15).
    Используй на вопросы: «кому отправить подборку», «кто давно не покупал», «кого реактивировать»,
    «у кого скоро повод/день рождения», «кому зайдёт тема X», «кому предложить VIP»."""
    if not tg_user_id:
        return "Не указан tg_user_id — не могу определить сотрудника."
    mode = (mode or "reactivate").strip().lower()

    try:
        if mode in ("reactivate", "vip"):
            data = call("crm-rfm.php", tg=str(tg_user_id))
            segs = data.get("segments", []) if isinstance(data, dict) else (data or [])
            by_key = {s.get("key"): s for s in segs if isinstance(s, dict)}
            keys = ("at_risk", "sleeping") if mode == "reactivate" else ("big_spenders",)
            pool = []
            for k in keys:
                pool.extend(by_key.get(k, {}).get("clients", []) or [])
            if not pool:
                return ("Нет подходящих клиентов в этих сегментах "
                        "(возможно, у тебя ещё мало закрытых сделок для RFM-анализа).")
            pool.sort(key=lambda c: float(c.get("total") or 0), reverse=True)
            # «Достижимые» = есть имя или телефон (кому реально можно отправить подборку).
            # Остальные — сироты: сделка есть, а карточки клиента с контактом нет.
            reachable = [c for c in pool if (c.get("name") or c.get("phone"))]
            orphan = len(pool) - len(reachable)
            title = ("🔄 РЕАКТИВАЦИЯ — ценные, но затихшие (отправить подборку, чтобы напомнить о себе):"
                     if mode == "reactivate"
                     else "💰 VIP-КАНДИДАТЫ — крупные разовые покупатели (персональный оффер):")
            lines = [title]
            if not reachable:
                lines.append("У всех кандидатов не заполнена карточка (нет имени и телефона) — "
                             "отправить подборку некому, сначала стоит завести контакты.")
            for c in reachable[:limit]:
                since = _days_since(c.get("last_purchase"))
                ago = f"молчит {since} дн." if since < 9999 else "дата покупки неизвестна"
                lines.append(f"• {_client_label(c)} — куплено на {_rub(c.get('total'))}, {ago}")
            more = len(reachable) - limit
            if more > 0:
                lines.append(f"…и ещё {more} с контактами. Скажи, если нужно больше или сделать подборку кому-то из списка.")
            if orphan > 0:
                lines.append(f"⚠️ Ещё {orphan} ценных сделок без карточки клиента (контакт не заполнен) — "
                             "их не охватить, пока не заведём данные.")
            return "\n".join(lines)

        clients = _scan_clients(tg_user_id)
        if not clients:
            return "Клиентов не нашлось (или у тебя пока пустая база)."

        if mode == "occasion":
            hits = []
            for c in clients:
                best = None
                for field, emoji in (("birthday", "🎂 ДР"), ("anniversary", "💍 годовщина"),
                                     ("name_day", "😇 именины")):
                    d = _days_until_event(c.get(field))
                    if d <= days and (best is None or d < best[0]):
                        best = (d, emoji, c.get(field))
                if best:
                    hits.append((best[0], best[1], _client_label(c), best[2]))
            if not hits:
                return (f"У клиентов нет ближайших поводов в горизонте {days} дн. "
                        "(даты ДР/годовщин/именин заполнены не у всех — можно расширить `days`).")
            hits.sort(key=lambda x: x[0])
            lines = [f"📅 СКОРО ПОВОД (в ближайшие {days} дн.) — подборка к дате:"]
            for d, emoji, label, raw in hits[:limit]:
                when = "сегодня" if d == 0 else ("завтра" if d == 1 else f"через {d} дн.")
                lines.append(f"• {label} — {emoji} {when}")
            more = len(hits) - limit
            if more > 0:
                lines.append(f"…и ещё {more}.")
            return "\n".join(lines)

        if mode == "theme":
            q = (query or "").strip().lower()
            if not q:
                return "Для mode=theme укажи тему в query (напр. query=\"Месси\" или \"бокс\")."
            hits = []
            for c in clients:
                hay_parts = []
                iv = c.get("interests")
                if isinstance(iv, list):
                    hay_parts.append(" ".join(str(x) for x in iv))
                elif iv:
                    hay_parts.append(str(iv))
                for f in ("tags", "dossier", "notes"):
                    if c.get(f):
                        hay_parts.append(str(c.get(f)))
                hay = " ".join(hay_parts).lower()
                if q in hay:
                    hits.append(_client_label(c))
            if not hits:
                return (f"Не нашёл клиентов с интересом «{query}» "
                        "(интересы/теги заполнены не у всех — стоит вести досье в карточках).")
            lines = [f"🎯 ТЕМА «{query}» — кому это близко (подборка под интерес):"]
            lines += [f"• {h}" for h in hits[:limit]]
            more = len(hits) - limit
            if more > 0:
                lines.append(f"…и ещё {more}.")
            return "\n".join(lines)

        return ("Неизвестный режим. Доступно: reactivate (затихшие ценные), "
                "vip (крупные разовые), occasion (скоро повод), theme (под тему/персону).")
    except Exception as e:
        return f"Не удалось собрать аналитику по клиентам: {e}"


@mcp.tool()
def deals_summary(tg_user_id: str, days: int = 30) -> str:
    """Сводка по ВСЕМ сделкам компании (только владелец/директор — сервер отдаёт
    все сделки только этим ролям; менеджеру вернутся лишь его собственные).
    days — за сколько последних дней считать (по умолчанию 30).
    Показывает: всего/по статусам, сумма выигранных, разбивка по менеджерам."""
    if not tg_user_id:
        return "Не указан tg_user_id."
    try:
        data = call("crm-deals.php", tg=str(tg_user_id))
    except Exception as e:
        return f"Ошибка доступа к сделкам: {e}"
    deals = data if isinstance(data, list) else data.get("deals", [])
    if not deals:
        return "Сделок нет."
    import datetime
    cutoff = (datetime.date.today() - datetime.timedelta(days=max(1, min(365, days)))).isoformat()
    recent = [d for d in deals if str(d.get("created_at") or "") >= cutoff]
    pool = recent or deals
    by_status, by_manager, won_sum = {}, {}, 0
    for d in pool:
        st = d.get("status") or "?"
        by_status[st] = by_status.get(st, 0) + 1
        m = d.get("assigned_to") or "—"
        try:
            amt = float(d.get("amount") or 0)
        except (TypeError, ValueError):
            amt = 0
        acc = by_manager.setdefault(m, {"n": 0, "won": 0.0})
        acc["n"] += 1
        if st == "won":
            acc["won"] += amt
            won_sum += amt
    def rub(v):
        return "{:,}".format(int(v)).replace(",", " ") + " ₽"
    lines = [f"Сделок за {days} дн.: {len(recent)} (всего в CRM: {len(deals)})"]
    lines.append("Статусы: " + ", ".join(f"{k}: {v}" for k, v in sorted(by_status.items(), key=lambda x: -x[1])))
    lines.append(f"Выиграно на сумму: {rub(won_sum)}")
    lines.append("По менеджерам:")
    for m, acc in sorted(by_manager.items(), key=lambda x: -x[1]["won"]):
        lines.append(f"• {m}: {acc['n']} сделок, выиграно {rub(acc['won'])}")
    return "\n".join(lines)


@mcp.tool()
def client_briefing(tg_user_id: str, query: str) -> str:
    """БРИФИНГ перед звонком/встречей: всё, что знаем о клиенте, одним ответом —
    карточка (контакты, скидка, лояльность), интересы/досье/заметки, важные даты,
    покупки (won-сделки и заказы с сайта), ОТКРЫТЫЕ сделки, суммарно потрачено.
    Используй на вопросы «что я знаю про X», «брифинг по клиенту», «готовлюсь к звонку с X».
    query — имя/телефон клиента. tg_user_id — Telegram id из контекста."""
    if not tg_user_id or not query:
        return "Нужны tg_user_id и query (имя/телефон клиента)."
    try:
        found = call("crm-clients.php", tg=str(tg_user_id), params={"search": query, "limit": "3"})
    except Exception as e:
        return f"Ошибка поиска клиента: {e}"
    rows = found if isinstance(found, list) else (found.get("clients") or found.get("data") or [])
    if not rows:
        return f"Клиент «{query}» не найден."
    if len(rows) > 1:
        opts = "; ".join(f"{r.get('full_name') or r.get('first_name') or '?'} ({r.get('phone') or 'без тел.'})" for r in rows)
        return f"Нашёл несколько: {opts}. Уточни, о ком речь."
    cid = rows[0]["id"]
    try:
        c = call("crm-clients.php", tg=str(tg_user_id), params={"id": str(cid)})
    except Exception as e:
        return f"Ошибка чтения карточки: {e}"
    if not isinstance(c, dict):
        return "Карточка клиента не читается."

    name = (c.get("full_name") or " ".join(x for x in (c.get("first_name"), c.get("last_name")) if x) or "?").strip()
    lines = [f"📋 БРИФИНГ: {name}"]
    contacts = [x for x in (c.get("phone"), c.get("whatsapp") and f"WA {c['whatsapp']}",
                            c.get("telegram") and f"@{str(c['telegram']).lstrip('@')}", c.get("email")) if x]
    if contacts:
        lines.append("Контакты: " + ", ".join(str(x) for x in contacts))
    meta = []
    if c.get("company"):
        meta.append(f"компания: {c['company']}" + (f" ({c['position']})" if c.get("position") else ""))
    if c.get("assigned_to"):
        meta.append(f"менеджер: {c['assigned_to']}")
    if c.get("source"):
        meta.append(f"источник: {c['source']}")
    if meta:
        lines.append("; ".join(meta))
    perks = []
    if float(c.get("discount_percent") or 0):
        perks.append(f"скидка {int(float(c['discount_percent']))}%")
    if c.get("loyalty_status"):
        perks.append(f"лояльность: {c['loyalty_status']}")
    ts = float(c.get("total_spent") or 0)
    if ts:
        perks.append(f"всего потрачено: {_rub(ts)}")
    if perks:
        lines.append("💳 " + "; ".join(perks))
    dates = [f"{emoji} {c[f]}" for f, emoji in (("birthday", "🎂 ДР"), ("anniversary", "💍 годовщина"), ("name_day", "😇 именины")) if c.get(f)]
    if dates:
        lines.append("Даты: " + ", ".join(dates))
    iv = c.get("interests")
    interests = ", ".join(iv) if isinstance(iv, list) and iv else (str(iv) if iv else "")
    if interests in ("[]", "{}", "null"):
        interests = ""
    if interests:
        lines.append(f"🎯 Интересы: {interests}")
    if c.get("tags"):
        lines.append(f"Теги: {c['tags']}")
    if c.get("dossier"):
        lines.append(f"📎 Досье: {str(c['dossier'])[:400]}")
    if c.get("notes"):
        lines.append(f"🗒 Заметки: {str(c['notes'])[:400]}")

    won = c.get("won_deals") or []
    if won:
        lines.append(f"\n✅ Покупки ({len(won)}):")
        for d in won[:6]:
            when = str(d.get("closed_at") or d.get("created_at") or "")[:10]
            lines.append(f"• {d.get('title') or '—'} — {_rub(d.get('amount'))} ({when})")
        if len(won) > 6:
            lines.append(f"…и ещё {len(won) - 6}")
    orders = c.get("site_orders") or c.get("orders") or []
    if orders:
        lines.append(f"🛒 Заказы с сайта: {len(orders)}")

    # Открытые сделки этого клиента (сервер отдаёт по роли; фильтруем по client_id).
    try:
        deals = call("crm-deals.php", tg=str(tg_user_id))
        deals = deals if isinstance(deals, list) else deals.get("deals", [])
        open_deals = [d for d in deals if str(d.get("client_id")) == str(cid)
                      and (d.get("status") or "") not in ("won", "lost")]
        if open_deals:
            lines.append(f"\n🔥 Открытые сделки ({len(open_deals)}):")
            for d in open_deals[:5]:
                lines.append(f"• {d.get('title') or '—'} — {d.get('status')} — {_rub(d.get('amount'))} (#{d.get('id')})")
    except Exception:
        pass

    won_total = sum(float(d.get("amount") or 0) for d in won)
    if won and not ts:
        lines.append(f"\nИтого покупок: {_rub(won_total)}")
    return "\n".join(lines)


@mcp.tool()
def client_find(tg_user_id: str, query: str, limit: int = 5) -> str:
    """Найти клиента в CRM по имени/телефону/email/telegram. Показывает карточку:
    контакты, город, персональную скидку, день рождения, закреплённого менеджера,
    заметки. id из ответа используй в deal_create/client_edit/client_note_add.
    """
    if not tg_user_id:
        return "Не указан tg_user_id."
    if not (query or "").strip():
        return "Скажи, кого искать (имя/телефон/email)."
    try:
        found = call("crm-clients.php", tg=str(tg_user_id), params={"search": query.strip()})
    except Exception as e:
        return f"Ошибка поиска клиентов: {e}"
    rows = found if isinstance(found, list) else (found or {}).get("clients", [])
    if not rows:
        return f"Клиент «{query}» в CRM не найден."
    out = []
    for c in rows[:max(1, min(10, limit))]:
        nm = (c.get("full_name")
              or " ".join(p for p in (c.get("first_name"), c.get("last_name")) if p)
              or c.get("company") or "?")
        bits = [f"• {nm} [id {c.get('id')}]"]
        contacts = ", ".join(str(c.get(k)) for k in ("phone", "telegram", "email") if c.get(k))
        if contacts:
            bits.append(f"  контакты: {contacts}")
        extra = []
        if c.get("city"):
            extra.append(c["city"])
        if c.get("discount_percent"):
            extra.append(f"скидка {c['discount_percent']}%")
        if c.get("birthday"):
            extra.append(f"ДР {c['birthday']}")
        if c.get("assigned_to"):
            extra.append(f"менеджер: {c['assigned_to']}")
        if extra:
            bits.append("  " + "; ".join(extra))
        if c.get("notes"):
            bits.append(f"  заметки: {str(c['notes'])[:150]}")
        out.append("\n".join(bits))
    return f"Найдено {len(rows)}:\n" + "\n".join(out)


@mcp.tool()
def client_create(tg_user_id: str, name: str, phone: str = "", telegram: str = "",
                  email: str = "", city: str = "", birthday: str = "",
                  notes: str = "", confirm: bool = False) -> str:
    """Создать НОВОГО клиента в CRM (закрепляется за написавшим менеджером).
    БЕЗ confirm=true — ПРЕДПРОСМОТР; с confirm=true — создаёт (после «да» менеджера).
    birthday — ГГГГ-ММ-ДД (если менеджер назвал). Перед созданием ПРОВЕРЬ дубль
    через client_find по телефону/имени.
    """
    if not tg_user_id:
        return "Не указан tg_user_id."
    if not (name or "").strip():
        return "Нужно имя клиента."
    manager = _manager_name(tg_user_id)
    if not confirm:
        parts = [f"ПРЕДПРОСМОТР нового клиента (ещё НЕ создан):", f"• Имя: {name}"]
        for label, v in (("Телефон", phone), ("Telegram", telegram), ("Email", email),
                         ("Город", city), ("ДР", birthday), ("Заметка", notes)):
            if v:
                parts.append(f"• {label}: {v}")
        parts.append(f"• Менеджер: {manager or '—'}")
        parts.append("Если верно — подтверди (client_create с confirm=true).")
        return "\n".join(parts)
    body = {"full_name": name.strip(), "created_by": manager, "assigned_to": manager}
    for key, v in (("phone", phone), ("telegram", telegram), ("email", email),
                   ("city", city), ("birthday", birthday), ("notes", notes)):
        if v:
            body[key] = v.strip()
    try:
        res = call("crm-clients.php", tg=str(tg_user_id), method="POST", body=body)
    except Exception as e:
        return f"Не удалось создать клиента: {e}"
    return f"Готово — клиент «{name}» создан (id {(res or {}).get('id', '?')}), закреплён за {manager}."


@mcp.tool()
def client_edit(tg_user_id: str, client_id: str, changes: dict, confirm: bool = False) -> str:
    """Правка карточки клиента. БЕЗ confirm=true — ПРЕДПРОСМОТР; с confirm=true — применяет.
    changes — словарь полей: phone, telegram, email, city, birthday (ГГГГ-ММ-ДД),
    anniversary, notes, discount_percent, interests, full_name и др.
    client_id — из client_find.
    """
    if not tg_user_id:
        return "Не указан tg_user_id."
    if not client_id or not isinstance(changes, dict) or not changes:
        return "Нужны client_id и непустой changes."
    if not confirm:
        pretty = ", ".join(f"{k} → {v}" for k, v in changes.items())
        return (f"ПРЕДПРОСМОТР: карточка клиента {client_id} изменится так: {pretty}.\n"
                f"Подтверди — применю (client_edit с confirm=true).")
    body = dict(changes)
    body["id"] = client_id
    try:
        call("crm-clients.php", tg=str(tg_user_id), method="PATCH", body=body)
    except Exception as e:
        return f"Не удалось изменить клиента: {e}"
    return f"Готово — карточка клиента {client_id} обновлена ({', '.join(changes)})."


@mcp.tool()
def client_note_add(tg_user_id: str, client_id: str, text: str) -> str:
    """Дописать заметку в карточку клиента (существующие заметки сохраняются,
    новая добавляется с датой). client_id — из client_find."""
    if not tg_user_id or not client_id or not (text or "").strip():
        return "Нужны client_id и текст заметки."
    manager = _manager_name(tg_user_id) or "бот"
    try:
        cur = call("crm-client-notes.php", tg=str(tg_user_id),
                   params={"client_id": str(client_id)})
    except Exception:
        cur = {}
    cur = cur if isinstance(cur, dict) else {}
    import datetime
    stamp = datetime.date.today().strftime("%d.%m.%Y")
    old = (cur.get("notes") or "").strip()
    merged = (old + "\n" if old else "") + f"[{stamp}, {manager}] {text.strip()}"
    body = {
        "client_id": str(client_id),
        "manager_id": cur.get("manager_id") or manager,
        "birthday": cur.get("birthday"),
        "preferences": cur.get("preferences"),
        "important_dates": cur.get("important_dates"),
        "notes": merged,
    }
    try:
        call("crm-client-notes.php", tg=str(tg_user_id), method="POST", body=body)
    except Exception as e:
        return f"Не удалось сохранить заметку: {e}"
    return f"Заметка добавлена клиенту {client_id}."


@mcp.tool()
def my_reminders(tg_user_id: str, upcoming_only: bool = True) -> str:
    """Напоминания написавшего менеджера (невыполненные). upcoming_only=true —
    только ближайшие 30 дней. Показывает id — им можно закрыть (reminder_done)."""
    if not tg_user_id:
        return "Не указан tg_user_id."
    manager = _manager_name(tg_user_id)
    if not manager:
        return "Не удалось определить менеджера."
    params = {"manager": manager}
    if upcoming_only:
        params["upcoming"] = "1"
    try:
        rows = call("crm-reminders.php", tg=str(tg_user_id), params=params)
    except Exception as e:
        return f"Ошибка чтения напоминаний: {e}"
    rows = rows if isinstance(rows, list) else []
    if not rows:
        return "Напоминаний нет."
    lines = []
    for r in rows[:20]:
        flag = "⚠️ ПРОСРОЧЕНО " if r.get("is_overdue") else ""
        lines.append(f"• {flag}{r.get('remind_at', '')} — {r.get('title', '')} (id {r.get('id')})")
    return f"Напоминания ({manager}):\n" + "\n".join(lines)


@mcp.tool()
def reminder_done(tg_user_id: str, reminder_id: str) -> str:
    """Отметить напоминание выполненным. reminder_id — из my_reminders."""
    if not tg_user_id or not reminder_id:
        return "Нужен reminder_id."
    try:
        call("crm-reminders.php", tg=str(tg_user_id),
             params={"id": str(reminder_id)}, method="PATCH", body={"is_done": True})
    except Exception as e:
        return f"Не удалось закрыть напоминание: {e}"
    return f"Напоминание {reminder_id} отмечено выполненным."


@mcp.tool()
def deals_at_risk(tg_user_id: str, limit: int = 10) -> str:
    """Сделки в ЗОНЕ РИСКА со скорингом и подсказкой следующего шага.

    Открытые сделки (не выиграны/проиграны), отсортированные по риску: застой
    (дней без обновления) + сумма + просроченная ожидаемая дата закрытия.
    Менеджеру — только его сделки; директору/owner/admin — по всей команде.
    Используй, когда спрашивают «что горит», «какие сделки в зоне риска»,
    «за какими сделками нужно следить», «что вот-вот сорвётся».
    tg_user_id — Telegram id написавшего (из контекста сессии).
    limit — сколько сделок вернуть (по умолчанию 10)."""
    if not tg_user_id:
        return "Не указан tg_user_id — не могу определить сотрудника."

    # Роль решает охват: менеджер видит свои, директор/owner/admin — все.
    my_name, role = "", "staff"
    try:
        who = call("bot-resolve-user.php", tg=str(tg_user_id), params={"tg": str(tg_user_id)})
        if isinstance(who, dict):
            my_name = who.get("name", "") or ""
            role = who.get("role", "staff") or "staff"
    except Exception:
        pass

    params = {"limit": str(max(1, min(50, int(limit))))}
    if role not in ("owner", "admin", "director") and my_name:
        params["manager"] = my_name

    try:
        data = call("crm-deals-at-risk.php", tg=str(tg_user_id), params=params)
    except Exception as e:
        return f"Ошибка доступа к сделкам: {e}"

    deals = data.get("deals", []) if isinstance(data, dict) else []
    if not deals:
        return "Сделок в зоне риска не найдено — воронка чистая."

    emoji = {"high": "🔴", "medium": "🟡", "low": "🟢"}
    lines = []
    for d in deals:
        amt = d.get("amount")
        try:
            amt_s = f"{int(float(amt)):,} ₽".replace(",", " ") if amt else "—"
        except (ValueError, TypeError):
            amt_s = "—"
        mark = emoji.get(d.get("risk_level"), "•")
        client = d.get("client_name") or "клиент не указан"
        title = d.get("title") or "—"
        stale = d.get("days_stale", 0)
        who_mgr = f" · {d.get('assigned_to')}" if role in ("owner", "admin", "director") and d.get("assigned_to") else ""
        lines.append(
            f"{mark} {client} — {title} — {amt_s} · {stale}д без движения{who_mgr} (#{d.get('id')})\n"
            f"   → {d.get('next_action', '')}"
        )

    header = "Сделки в зоне риска" + ("" if role in ("owner", "admin", "director") else f" ({my_name})") + f", {len(deals)}:"
    return header + "\n" + "\n".join(lines)




# ── Сделки из WhatsApp-чата (конвейер deal-capture, Ф2: диалог с Доком) ──

@mcp.tool()
def chat_intake_pending(tg_user_id: str) -> str:
    """Кандидаты в сделки, распознанные из WhatsApp-чата «STARGIFT⭐️», которые
    ждут решения этого менеджера (или ничейные). Показывай менеджеру кратко:
    экспонат, сумма, оплата, клиент; спрашивай «заношу?»."""
    if not str(tg_user_id or "").strip().isdigit():
        return "Нужен числовой tg_user_id из контекста сессии."
    try:
        data = call("chat-intake.php", tg=str(tg_user_id))
    except Exception as e:
        return f"Ошибка запроса ленты: {e}"
    rows = data.get("proposals") or []
    if not rows:
        return "Кандидатов из чата нет — всё разобрано."
    lines = []
    for r in rows[:10]:
        p = r.get("proposal") or {}
        items = " + ".join(i.get("name", "") for i in (p.get("items") or []) if i.get("name"))
        amt = p.get("amount")
        amt_s = f"{int(float(amt)):,} ₽".replace(",", " ") if amt else "сумма не распознана"
        client = (p.get("client") or {}).get("name") or p.get("client_hint") or "клиент не указан"
        lines.append(f"#{r['id']} · {items or '—'} · {amt_s} · {client} · «{(r.get('raw_text') or '')[:80]}»")
        for it in (p.get("items") or []):
            for c in (it.get("catalog") or [])[:3]:
                pid = c.get("product_id")
                price = c.get("price")
                price_s = f", на сайте {int(float(price)):,} ₽".replace(",", " ") if price else ""
                if pid:
                    lines.append(f"   ↳ каталог: {c.get('name','')[:60]}{price_s} — https://stargift.ru/product/{pid}/")
    return "Кандидаты из чата (номер — для подтверждения):\n" + "\n".join(lines)


@mcp.tool()
def chat_intake_confirm(tg_user_id: str, intake_id: int, amount: int = 0,
                        client_phone: str = "", client_name: str = "",
                        client_notes: str = "", client_interests: str = "",
                        title: str = "", payment: str = "",
                        product_id: str = "", gallery: str = "",
                        supplies: list = None, assigned_to: str = "",
                        notes: str = "", handled_by: str = "",
                        confirm: bool = False) -> str:
    """Занести кандидата #intake_id как сделку (won) в CRM. ОБЯЗАТЕЛЬНО сначала
    покажи менеджеру, что именно будет занесено, и получи явное «да» —
    только после этого вызывай с confirm=True. Недостающие данные (телефон
    клиента, уточнённая сумма) передавай параметрами. payment: нал/карта/
    перевод/счёт/ссылка/сбп. client_notes — как познакомились/детали сделки,
    client_interests — интересы клиента (попадут в карточку клиента в CRM).
    product_id — подтверждённый менеджером экспонат каталога (будет скрыт с сайта).
    gallery — где продано (Времена года / Гименей / Гименей склад / Гименей Прайм /
    DreamHouse). supplies — список названий расходников для списания со склада
    ("Подставка Футбол — Чёрный мрамор", "Акрил — Футбол", "Книжный короб — Средний",
    "Чехлы гитары — Акустика"...); проверяй варианты через supply_options.
    assigned_to — если сделка ДРУГОГО менеджера (сказали «занеси на X»): ТОЧНОЕ имя
    из CRM («Анна Коротеева»); без этого параметра сделка запишется на подтверждающего.
    handled_by — КТО ВЁЛ сделку, если она канальная: продажа с сайта/Instagram →
    assigned_to='Сайт/Instagram', handled_by='<имя человека>' (напр. «Анна Поповская»).
    notes — комментарий в саму сделку (например, длинный фидбэк менеджера из чата);
    НЕ путай с client_notes (те идут в карточку клиента). Если попросили «запиши
    комментарий в сделку», а параметр не передан — сделка останется БЕЗ комментария:
    никогда не отчитывайся о том, чего не передавал."""
    if not confirm:
        return "Нужно подтверждение менеджера. Покажи детали и вызови с confirm=True после явного «да»."
    if not str(tg_user_id or "").strip().isdigit():
        return "Нужен числовой tg_user_id из контекста сессии."
    edits = {}
    if amount:
        edits["amount"] = amount
    if client_phone:
        edits["client_phone"] = client_phone
    if client_name:
        edits["client_name"] = client_name
    if client_notes:
        edits["client_notes"] = client_notes
    if client_interests:
        edits["client_interests"] = client_interests
    if product_id:
        edits["product_id"] = str(product_id)
    if gallery:
        edits["gallery"] = gallery
    if supplies:
        edits["supplies"] = [str(x) for x in supplies if str(x).strip()]
    if title:
        edits["title"] = title
    if payment:
        edits["payment"] = payment
    if assigned_to:
        edits["assigned_to"] = assigned_to
    if handled_by:
        edits["handled_by"] = handled_by
    if notes:
        edits["notes"] = notes
    try:
        res = call("chat-intake.php", tg=str(tg_user_id), method="POST",
                   body={"intake_id": int(intake_id), "decision": "approve", "edits": edits})
    except Exception as e:
        return f"Ошибка занесения: {e}"
    if res.get("ok"):
        stock = res.get("stock_written_off", 0)
        parts = [f"✅ Занесено (deal {res.get('deal_id')}, {res.get('action')})."]
        if assigned_to:
            parts.append(f"Ответственный менеджер: {assigned_to}.")
        parts.append("Комментарий сохранён в сделку." if notes
                     else "Комментарий в сделку НЕ передавался.")
        parts.append(f"Экспонат каталога: скрыт с сайта ({stock} поз.)." if stock
                     else "Экспонат каталога: не привязан — с сайта ничего не скрыто.")
        for sr in (res.get("supplies") or []):
            if sr.get("ok"):
                parts.append(f"Расходник списан: {sr['name']} ({sr.get('location','')}, остаток {sr.get('left')}).")
            else:
                parts.append(f"⚠️ Расходник НЕ списан: {sr['name']} — {sr.get('err')}.")
        return " ".join(parts)
    return f"Не получилось: {res.get('error', res)}"


@mcp.tool()
def chat_intake_reject(tg_user_id: str, intake_id: int, reason: str = "") -> str:
    """Отклонить кандидата #intake_id (это не сделка / дубль / шутка в чате).
    Вызывай после явного «нет» менеджера."""
    if not str(tg_user_id or "").strip().isdigit():
        return "Нужен числовой tg_user_id из контекста сессии."
    try:
        res = call("chat-intake.php", tg=str(tg_user_id), method="POST",
                   body={"intake_id": int(intake_id), "decision": "reject"})
    except Exception as e:
        return f"Ошибка: {e}"
    return "❌ Отклонено." if res.get("ok") else f"Не получилось: {res.get('error', res)}"




@mcp.tool()
def supply_options(query: str = "") -> str:
    """Справочник расходников склада с остатками по локациям. query — фильтр по
    названию/категории (например "подставка футбол", "короб", "акрил", "чехол").
    Используй, чтобы предложить менеджеру точные варианты перед списанием."""
    try:
        data = call("supply-stock.php", tg="311452113")
    except Exception as e:
        return f"Ошибка склада: {e}"
    rows = data if isinstance(data, list) else data.get("rows") or data.get("stock") or []
    q = (query or "").lower().split()
    agg = {}
    for r in rows:
        name = r.get("name") or r.get("supply_name") or ""
        cat = r.get("category") or r.get("supply_category") or ""
        loc = r.get("location") or ""
        qty = int(float(r.get("quantity") or 0))
        hay = (name + " " + cat).lower()
        if q and not all(w in hay for w in q):
            continue
        key = name
        agg.setdefault(key, {"cat": cat, "locs": {}})
        if qty > 0:
            agg[key]["locs"][loc] = agg[key]["locs"].get(loc, 0) + qty
    if not agg:
        return "Ничего не нашлось. Попробуй другой запрос (подставка/акрил/короб/чехол/коробка)."
    lines = []
    for name, info in sorted(agg.items())[:25]:
        locs = ", ".join(f"{l}: {n}" for l, n in sorted(info["locs"].items())) or "нет остатков"
        lines.append(f"• {name} — {locs}")
    return "Расходники (остатки по локациям):\n" + "\n".join(lines)

@mcp.tool()
def make_price_tag(tg_user_id: str, product_query: str = "", name: str = "",
                   subtitle: str = "", price: int = -1, kind: str = "autograph",
                   with_photo: bool = True, photo: str = "") -> str:
    """Сгенерировать ЦЕННИК экспоната (фирменный дизайн CRM, 150×100 мм, 300 dpi —
    можно сразу в печать) и прислать PNG в этот чат.
    product_query — экспонат из каталога («Тайсон перчатка»): имя, подзаголовок,
    цена и фото возьмутся из карточки автоматически. ИМЯ ВСЕГДА бери ЦЕЛИКОМ из
    карточки каталога — включая группу/уточнение в скобках: «Крист Новоселич
    (Nirvana)», НЕ «Крист Новоселич» (правило Вашика 22.07). name/subtitle/price —
    переопределять ТОЛЬКО если менеджер явно попросил другой текст
    (price=0 — ценник БЕЗ цены, для лотов «по запросу»).
    kind: autograph | book (книжный: узкий трекинг; заголовок всегда капсом).
    with_photo=False — текстовый ценник без фото.
    photo — заменить фото: https-URL ИЛИ локальный путь. Если менеджер ПРИСЛАЛ
    фото в чат — возьми путь этого файла из сообщения и передай сюда.
    ПРАВКИ ИЗ ЧАТА («замени цену на 300к», «без цены», «другое фото», «возьми
    фото из чата») — просто вызови тул ЗАНОВО с изменённым параметром, остальное
    снова подтянется из карточки."""
    if not str(tg_user_id or "").strip().isdigit():
        return "Нужен числовой tg_user_id из контекста сессии."
    import price_tag as _pt

    p = None
    if product_query:
        found = _find_products(product_query, 3)
        if not found:
            return f"В каталоге не нашёл «{product_query}» — уточни название или задай name/price вручную."
        p = found[0]
        name = name or (p.get("title") or "")
        subtitle = subtitle or (p.get("description") or "")
        if price < 0:
            price = int(float(p.get("price") or 0))
        if not photo and with_photo:
            photos = p.get("photos") or []
            photo = photos[0] if photos else ""
    if not with_photo:
        photo = ""
    if not name:
        return "Нужен product_query или name."
    if price < 0:
        price = 0
    try:
        png = _pt.render_price_tag(name, subtitle, price, photo_url=photo, kind=kind)
    except Exception as e:
        return f"Ошибка рендера ценника: {e}"
    cap = f"Ценник: {name}" + (f" · {_fmt_price(price)}" if price else " · без цены")
    if _pt.send_png_telegram(str(tg_user_id), png, cap):
        # Вдогонку — тот же PNG документом: фото Telegram сжимает, печать — только из файла
        _pt.send_document_telegram(str(tg_user_id), png, "Файл для печати (150×100 мм, 300 dpi)",
                                   filename=f"Ценник — {name}.png")
        matched = f" (экспонат: {p.get('title')})" if p else ""
        return f"✅ Ценник отправлен: превью + файл для печати{matched}."
    return "Ценник срендерился, но отправка в Telegram не удалась."


@mcp.tool()
def make_certificate(tg_user_id: str, product_query: str = "", signed_by: str = "",
                     exhibit: str = "", photos: list = None) -> str:
    """Сгенерировать СЕРТИФИКАТ ПРОИСХОЖДЕНИЯ (A4, точная копия встроенного
    генератора CRM: водяной знак, золотая рамка, лого, «Сертификат происхождения /
    Certificate of authenticity», до 2 фото экспоната, «Подписано: …» /
    «Экспонат: …», футер с телефоном) и прислать PNG в этот чат.
    product_query — экспонат из каталога: «Подписано» = персона, «Экспонат» =
    тип из карточки, фото = первые 2 из карточки (экспонат + автограф крупно).
    signed_by / exhibit — переопределить строки вручную.
    photos — заменить фото: список https-URL или ЛОКАЛЬНЫХ путей (фото,
    присланные менеджером в чат — возьми пути из сообщений).
    ПРАВКИ ИЗ ЧАТА («замени фото», «подпиши иначе») — вызови тул заново с
    изменённым параметром."""
    if not str(tg_user_id or "").strip().isdigit():
        return "Нужен числовой tg_user_id из контекста сессии."
    import price_tag as _pt

    p = None
    if product_query:
        found = _find_products(product_query, 3)
        if not found:
            return f"В каталоге не нашёл «{product_query}»."
        p = found[0]
        signed_by = signed_by or (p.get("title") or "")
        exhibit = exhibit or (p.get("description") or "")
        if not photos:
            photos = (p.get("photos") or [])[:2]
    if not (signed_by or exhibit):
        return "Нужен product_query или signed_by/exhibit."
    try:
        png = _pt.render_certificate_a4(signed_by, exhibit, photo_urls=photos or [])
    except Exception as e:
        return f"Ошибка рендера сертификата: {e}"
    cap = f"Сертификат происхождения · {signed_by or exhibit}"
    if _pt.send_png_telegram(str(tg_user_id), png, cap):
        _pt.send_document_telegram(str(tg_user_id), png, "Файл для печати (A4, 300 dpi)",
                                   filename=f"Сертификат — {signed_by or exhibit}.png")
        return (f"✅ Сертификат отправлен: превью + файл для печати ({signed_by or exhibit}; "
                f"фото: {len(photos or [])}).")
    return "Сертификат срендерился, но отправка в Telegram не удалась."




@mcp.tool()
def make_presentation_premium(tg_user_id: str, title: str, csv_text: str = "",
                              items: list = None, kicker: str = "",
                              subtitle: str = "", cover_url: str = "",
                              sections: list = None, cover_credit: str = "",
                              omit_slides: list = None,
                              manager_name: str = "", manager_contact: str = "") -> str:
    """Собрать ПРЕМИАЛЬНУЮ презентацию по эталону v2 («Культовое кино»: 1920×1080,
    тёмный люкс, Prata + Golos Text) и прислать PDF в Telegram.

    Когда использовать: просят «премиальную презу / презентацию как культовое кино /
    презу по эталону / кинематографичную подборку». Обычная быстрая подборка по
    фирменному шаблону — это ДРУГОЙ инструмент, make_selection_pdf (он остаётся).

    Источники лотов (можно комбинировать):
    - csv_text="last" — менеджер прислал боту CSV-выгрузку сайта файлом
      (";", utf-8-sig: Название; Краткое описание; Цена; В наличии; URL на сайте;
      URL главного фото). Бот возьмёт последний присланный .csv.
    - photo_url лота принимает также "last" (последнее присланное боту фото —
      когда менеджер шлёт свой снимок экспоната) и локальный путь к файлу.
    - csv_text="<содержимое csv>" — если текст выгрузки уже в диалоге.
    - items — список позиций, которые ты собрал сам через catalog_search/catalog_get:
      {name, person, type, desc, price, cert, photo_url, group, instock, url}.
      group = раздел презентации (обычно подписант). person = имя подписанта
      (крупная строка подписи), type = краткий тип («Футболка с автографом»,
      «Мяч Nike с автографом») — подзаголовок под именем; подставки и витрины
      в type НЕ упоминать. desc: ПО УМОЛЧАНИЮ ПУСТАЯ СТРОКА. Перчатка, футболка,
      мяч, фото, шорты «с автографом» — самоочевидны, desc="" ВСЕГДА. Заполняй
      desc максимум у 2–3 лотов на ВСЮ подборку — только когда без него не понять
      предмет (что за брошюра, чем важен именно этот матч) — и не длиннее одного
      предложения (~120 знаков). Заполненный desc у каждого лота = ошибка,
      подборку вернут на переделку. desc НЕ должен пересказывать
      name/type экспоната — только новые факты: история, титулы, кто подписал,
      чем уникален.
    - sections — необязательный список разделов [{title, lead, epithet}]:
      lead — 1–2 фразы шмуцтитула, epithet — кикер для маленького раздела
      («ЭКСПОНАТЫ С АВТОГРАФОМ ЛЕГЕНДЫ СБОРНОЙ ПОРТУГАЛИИ»). title должен
      совпадать с group лотов.

    РЕГИСТР ТЕКСТА — правила дома как в make_selection_pdf: люкс-тон, без шуток и
    восклицаний. ⛔ ЗАПРЕТ темы подлинности: не писать «подлинный/подлинность/
    сертификат подлинности»; сертификатор — только спецификацией (JSA, Beckett).
    ⛔ Также запрещены (Вашик, 22.07): «реликвия», «под ключ», «частное собрание», «собирательный
    экспонат», «сопровождён спецификацией эксперта», механические и канцелярские обороты
    («снимок фиксирует», «раздел собирает», «предметы, связанные с»).
    ⛔ Слова «предмет(ы)» и «лот(ы)» в текстах запрещены — ВСЕГДА «экспонат(ы)» (Вашик, 23.07).
    Имена всегда полные; термины раскрывать; меньше тире;
    размеры не в desc; desc только если добавляет контекст (иначе пустой);
    факты не выдумывать; без буллет-пойнтов; не дублировать
    видимое на слайде («в наличии», «в раме» из названия).

    title — название подборки («Большая игра»), kicker — надзаголовок капсом
    («АВТОГРАФЫ ЛЕГЕНД ГОЛЬФА», «ЛЕГЕНДЫ СПОРТА»), subtitle — 2–3 предложения
    для обложки, cover_url — фото обложки (иначе возьмётся фото первого лота).
    ОБЛОЖКА: лучший вариант — имиджевый архивный кадр по теме из открытых
    источников: найди через archive_photo_search и передай cover_url=img
    + cover_credit=credit из результата (подпись источника обязательна).
    Если подходящего кадра нет — эффектное фото лота, cover_credit пустой.

    omit_slides — СПИСОК служебных слайдов, которые НЕ включать. Служебные слайды
    добавляются автоматически, и убрать их можно ТОЛЬКО этим параметром (заново
    пересобрав презентацию) — «удалить слайд N» иначе не сработает. Значения:
    "contents" (Состав собрания), "instock" (Экспонаты в наличии),
    "index" (Полный перечень), "contacts" (последний слайд с контактами).
    Значения: "contents", "instock", "index", "contacts", "framing" (Примеры
    оформления), "delivery" (Доставка и подарочный сертификат).
    Пример: менеджер просит «убери два последних слайда» → сверься со slides_summary
    из get_last_presentation. Слайды «framing» и «delivery» идут перед contacts.

    manager_name / manager_contact — персональный менеджер клиента: имя и контакт
    (телефон/@telegram). Показываются отдельной колонкой на слайде контактов —
    делает общение личным. Если менеджер не указан — колонки просто нет.
    По умолчанию в конце добавляются слайды «Примеры оформления» (как экспонат
    выглядит в раме) и «Доставка и подарочный сертификат» — убрать через omit_slides.

    Сборка идёт из единых шаблонов конструктора подборок (эталон v2) — правки
    шаблонов автоматически подхватываются. Рендер занимает ~20–40 секунд.
    """
    if not tg_user_id:
        return "Не указан tg_user_id."
    if not title:
        return "Передай title — название подборки."
    import presentation_v2
    return presentation_v2.make_and_send(str(tg_user_id), title, csv_text=csv_text,
                                         items=items, kicker=kicker,
                                         subtitle=subtitle, cover_url=cover_url,
                                         sections=sections, cover_credit=cover_credit,
                                         omit_slides=omit_slides,
                                         manager_name=manager_name,
                                         manager_contact=manager_contact)


@mcp.tool()
def archive_photo_search(query: str, limit: int = 8) -> str:
    """Поиск ИМИДЖЕВЫХ фото в открытых источниках — для обложек и атмосферных
    слайдов премиальных презентаций. Ищет сразу в Wikimedia Commons И Openverse
    (агрегатор Flickr, музеев, госархивов). Только лицензии, разрешающие
    коммерческое использование: public domain / CC0 / CC BY / CC BY-SA
    (NC и ND отфильтрованы). Работает и для СОВРЕМЕННЫХ матчей/персон — на
    Flickr много живой спортивной съёмки под CC.

    query — англоязычный запрос («Muhammad Ali boxing 1966», «Zlatan Ibrahimovic
    match», «vintage football stadium crowd»).

    ПОДПИСЬ КАДРА (правка Вашика 27.07): лицензии и «Wikimedia Commons» на слайдах
    НЕ пишем. cover_credit формулируй сам по title кадра — коротко кто/где/когда
    изображён по-русски («Пеле после финала чемпионата мира, стадион «Ацтека», 1970»).
    Исключение: для лицензий CC BY / CC BY-SA добавь в конец « · фото: <автор>» —
    это юридическое требование атрибуции; для public domain ничего не добавляем.

    Возвращает JSON {results, licensed_search}:
    - results: [{img, page, title, license, author, source, width, credit}] —
      img → cover_url/photo_url, credit → cover_credit (готовая подпись, не менять),
      page — страница источника (можно дать менеджеру для проверки).
    - licensed_search: ссылки на поиск в ПЛАТНЫХ фотобанках (ТАСС, РИА Новости,
      Getty) по этому же запросу. Если свободного кадра нет или менеджеру нужен
      конкретный кадр матча/фильма — дай эти ссылки: менеджер сам выберет и
      купит лицензию, ставить такие фото в презентацию без покупки НЕЛЬЗЯ.
    ⛔ Кадры/постеры современных фильмов свободными не бывают — не подставляй
    похожие случайные фото; для кино: фото НАШЕГО экспоната или licensed_search.
    Генерацией имиджевые кадры не делаем (решение Вашика) — если ничего не
    нашлось, честно скажи и предложи licensed_search."""
    import json as _json
    import urllib.request as _rq
    import urllib.parse as _up
    import re as _re
    q = str(query or "").strip()
    if len(q) < 2:
        return "Передай query — тема поиска (по-английски)."
    lim = max(1, min(int(limit or 8), 16))
    strip = _re.compile(r"<[^>]+>")
    items = []

    # 1) Wikimedia Commons
    try:
        api = "https://commons.wikimedia.org/w/api.php?" + _up.urlencode({
            "action": "query", "format": "json", "generator": "search",
            "gsrsearch": "filetype:bitmap " + q, "gsrnamespace": 6, "gsrlimit": 24,
            "prop": "imageinfo", "iiprop": "url|extmetadata|size", "iiurlwidth": 1600,
        })
        req = _rq.Request(api, headers={"User-Agent": "StargiftPodborki/1.0 (internal tool)"})
        data = _json.loads(_rq.urlopen(req, timeout=40).read())
        free = _re.compile(r"public\s*domain|pd-|cc0|cc[- ]by(?:[- ]sa)?[- 0-9.]*$", _re.I)
        for p in (data.get("query", {}).get("pages", {}) or {}).values():
            ii = (p.get("imageinfo") or [None])[0]
            if not ii:
                continue
            meta = ii.get("extmetadata") or {}
            lic = strip.sub("", (meta.get("LicenseShortName") or {}).get("value") or "").strip()
            if not free.search(lic) or int(ii.get("width") or 0) < 900:
                continue
            author = strip.sub("", (meta.get("Artist") or {}).get("value") or "").strip()
            title = _re.sub(r"^File:|\.[a-z]+$", "", p.get("title") or "", flags=_re.I)
            items.append({
                "img": ii.get("thumburl") or ii.get("url"),
                "page": ii.get("descriptionurl") or "",
                "title": title, "license": lic, "author": author[:80],
                "source": "Wikimedia Commons", "width": int(ii.get("width") or 0),
                "credit": "архив: Wikimedia Commons · " + lic.lower()
                          + (" · " + author[:60] if author else ""),
            })
    except Exception:
        pass

    # 2) Openverse (Flickr, музеи, госархивы) — только коммерчески чистые лицензии
    try:
        api = "https://api.openverse.org/v1/images/?" + _up.urlencode({
            "q": q, "license": "cc0,pdm,by,by-sa", "page_size": 20})
        req = _rq.Request(api, headers={"User-Agent": "StargiftPodborki/1.0 (internal tool)"})
        data = _json.loads(_rq.urlopen(req, timeout=40).read())
        for r in data.get("results") or []:
            w = int(r.get("width") or 0)
            if w and w < 900:
                continue
            lic = (r.get("license") or "").upper()
            lic = "Public domain" if lic in ("CC0", "PDM") else "CC " + lic.replace("-", " ")
            src = (r.get("source") or "openverse").replace("_", " ").title()
            author = (r.get("creator") or "").strip()
            items.append({
                "img": r.get("url"), "page": r.get("foreign_landing_url") or "",
                "title": (r.get("title") or "")[:100], "license": lic,
                "author": author[:80], "source": src, "width": w,
                "credit": "архив: " + src + " · " + lic.lower()
                          + (" · " + author[:60] if author else ""),
            })
    except Exception:
        pass

    # дедуп по img, крупные вперёд
    seen, dedup = set(), []
    for it in sorted(items, key=lambda x: -(x.get("width") or 0)):
        if it.get("img") and it["img"] not in seen:
            seen.add(it["img"])
            dedup.append(it)
    qq = _up.quote(q)
    out = {
        "results": dedup[:lim],
        "licensed_search": {
            "ТАСС": "https://tassphoto.com/ru/search?query=" + qq,
            "РИА Новости": "https://visualrian.ru/search/?query=" + qq,
            "Getty Images": "https://www.gettyimages.com/photos/" + qq,
        },
    }
    if not dedup:
        out["note"] = ("Свободных кадров не нашлось. Предложи менеджеру licensed_search — "
                       "выбрать и лицензировать кадр самостоятельно.")
    return _json.dumps(out, ensure_ascii=False)


@mcp.tool()
def get_last_presentation(tg_user_id: str) -> str:
    """Данные ПОСЛЕДНЕЙ премиальной презентации этого чата (make_presentation_premium
    сохраняет их автоматически). Используй, когда просят ПОПРАВИТЬ уже присланную
    подборку («поправь слайд 4», «замени фото у Роналду», «убери третий лот») —
    файл презентации заново просить НЕ нужно.

    Возвращает JSON: theme, sections, lots (полные поля каждого лота),
    slides_summary — карту слайдов (номер, тип, заголовок, лоты) для сопоставления
    «слайд №N» с содержимым, и omit_slides — уже убранные служебные слайды.
    Как править: измени нужные lots/sections/theme и вызови make_presentation_premium
    заново с items=исправленные lots (поля person/type/desc/price/cert/img→photo_url/
    group/instock/url сохраняй), sections, той же темой И тем же omit_slides (плюс
    новые, если просят убрать ещё слайды) — придёт обновлённый PDF + Keynote.
    УБРАТЬ СЛАЙД можно только через omit_slides (см. make_presentation_premium):
    служебные слайды нельзя «удалить» иначе — только не включить при пересборке."""
    if not str(tg_user_id or "").strip():
        return "Не указан tg_user_id."
    import presentation_v2
    state = presentation_v2.load_last_state(str(tg_user_id))
    if not state:
        return ("Сохранённой презентации для этого чата нет — попроси файл/CSV "
                "или собери подборку заново.")
    import json as _json
    return _json.dumps(state, ensure_ascii=False)


@mcp.tool()
def make_print_files(tg_user_id: str, price_tags: list = None,
                     certificates: list = None) -> str:
    """ПАЧКА печатных файлов одним ZIP-архивом — когда менеджер просит СРАЗУ
    МНОГО ценников и/или сертификатов («сделай ценники на все новинки спорта»).
    price_tags — список запросов экспонатов для ценников (каждый — как
    product_query в make_price_tag: «Тайсон перчатка», «Новоселич гитара»…).
    certificates — то же для сертификатов A4.
    1-2 файла суммарно — пришлёт файлами без архива; 3+ — одним ZIP для скачки.
    Имена в архиве: «Ценник — <имя из каталога>.png» / «Сертификат — <имя>.png».
    Не найденные экспонаты честно перечисляются в ответе."""
    if not str(tg_user_id or "").strip().isdigit():
        return "Нужен числовой tg_user_id из контекста сессии."
    import price_tag as _pt

    files, missed = [], []
    for q in (price_tags or []):
        q = str(q).strip()
        if not q:
            continue
        found = _find_products(q, 3)
        if not found:
            missed.append(f"ценник: {q}")
            continue
        p = found[0]
        name = p.get("title") or q
        photos = p.get("photos") or []
        try:
            png = _pt.render_price_tag(name, p.get("description") or "",
                                       int(float(p.get("price") or 0)),
                                       photo_url=photos[0] if photos else "")
            files.append((png, f"Ценник — {name}.png"))
        except Exception as e:
            missed.append(f"ценник: {q} (рендер: {e})")
    for q in (certificates or []):
        q = str(q).strip()
        if not q:
            continue
        found = _find_products(q, 3)
        if not found:
            missed.append(f"сертификат: {q}")
            continue
        p = found[0]
        name = p.get("title") or q
        try:
            png = _pt.render_certificate_a4(name, p.get("description") or "",
                                            photo_urls=(p.get("photos") or [])[:2])
            files.append((png, f"Сертификат — {name}.png"))
        except Exception as e:
            missed.append(f"сертификат: {q} (рендер: {e})")

    if not files:
        return "Ничего не срендерилось." + (" Не найдено: " + "; ".join(missed) if missed else "")

    miss_note = (" ⚠️ Не найдено/не вышло: " + "; ".join(missed)) if missed else ""
    if len(files) <= 2:
        sent = sum(1 for path, fname in files
                   if _pt.send_document_telegram(str(tg_user_id), path,
                                                 "Файл для печати (300 dpi)", filename=fname))
        return f"✅ Отправлено файлами: {sent} из {len(files)}.{miss_note}"
    zip_path = _pt.make_zip(files)
    ok = _pt.send_document_telegram(str(tg_user_id), zip_path,
                                    f"Печатные файлы: {len(files)} шт (300 dpi)",
                                    filename="stargift-print.zip")
    if ok:
        return f"✅ ZIP с {len(files)} печатными файлами отправлен.{miss_note}"
    return f"ZIP собран, но отправка не удалась.{miss_note}"


@mcp.tool()
def send_file(tg_user_id: str, filename: str, text_content: str = "",
              rows: list = None, sheet_name: str = "Лист1",
              source_path: str = "") -> str:
    """Создать ФАЙЛ и прислать его в чат документом — для любых просьб менеджеров
    «переведи в другой формат», «сделай таблицу/список файлом», «выгрузи в CSV/Excel».
    filename — имя с расширением: .csv / .txt / .md / .json / .html → передай
    содержимое в text_content (для CSV — строки через \\n, значения через запятую
    или ;). .xlsx → передай rows: список строк-списков (первая — заголовки),
    text_content не нужен. Данные бери из read_document/каталога/CRM или из слов
    менеджера. Файл уходит без сжатия, открывается в Excel/Numbers.
    «СКИНЬ ФАЙЛОМ / в полном качестве / без сжатия» про УЖЕ СУЩЕСТВУЮЩИЙ файл
    (генерация оформления, PDF, присланное фото) → source_path = путь файла
    (из результата exhibit_photo_frame; "last" = последняя генерация оформления);
    text_content и rows не нужны, filename можно опустить."""
    if not str(tg_user_id or "").strip().isdigit():
        return "Нужен числовой tg_user_id из контекста сессии."
    import tempfile as _tf
    from pathlib import Path as _P
    if (source_path or "").strip():
        sp_raw = source_path.strip()
        if sp_raw == "last":
            import framing_tool as _ft
            sp_raw = _ft.latest_render() or ""
            if not sp_raw:
                return "Генераций оформления пока нет."
        sp = _P(sp_raw)
        _ALLOWED_ROOTS = ("/Users/docbrown/stargift-framing-tests/",
                          "/Users/docbrown/hermes-doc/premium-decks/",
                          "/Users/docbrown/.hermes/profiles/staff/image_cache/",
                          _tf.gettempdir())
        if not sp.exists():
            return f"Файл не найден: {sp}"
        if not any(str(sp).startswith(r) for r in _ALLOWED_ROOTS):
            return "Этот файл отправить нельзя (вне рабочих папок)."
        import urllib.request as _rq, json as _json, uuid as _uuid
        from selection_sender import _telegram_token as _tok
        token = _tok()
        boundary = _uuid.uuid4().hex
        fname = (filename or sp.name).replace("/", "_")
        crlf = "\r\n"
        body = (f"--{boundary}{crlf}Content-Disposition: form-data; name=\"chat_id\"{crlf}{crlf}{tg_user_id}{crlf}").encode()
        body += (f"--{boundary}{crlf}Content-Disposition: form-data; name=\"document\"; "
                 f"filename=\"{fname}\"{crlf}Content-Type: application/octet-stream{crlf}{crlf}").encode()
        body += sp.read_bytes() + f"{crlf}--{boundary}--{crlf}".encode()
        req = _rq.Request(f"https://api.telegram.org/bot{token}/sendDocument", data=body,
                          headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
        try:
            ok = _json.loads(_rq.urlopen(req, timeout=120).read()).get("ok")
        except Exception as e:
            return f"Не отправилось: {e}"
        return f"Файл «{fname}» отправлен документом (без сжатия)." if ok else "Telegram отказал в отправке."
    fn = (filename or "file.txt").strip().replace("/", "_")
    workdir = _P(_tf.mkdtemp(prefix="sg-file-"))
    path = workdir / fn
    ext = path.suffix.lower()
    try:
        if ext == ".xlsx":
            if not rows:
                return "Для .xlsx нужен rows — список строк (первая строка = заголовки)."
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = (sheet_name or "Лист1")[:30]
            for r in rows:
                ws.append(["" if c is None else c for c in (r if isinstance(r, list) else [r])])
            wb.save(str(path))
        else:
            if not text_content:
                return "Нужен text_content (или rows для .xlsx)."
            enc = "utf-8-sig" if ext == ".csv" else "utf-8"  # BOM — чтобы Excel понял кириллицу
            path.write_text(text_content, encoding=enc)
    except Exception as e:
        return f"Не удалось создать файл: {e}"
    import price_tag as _pt
    if _pt.send_document_telegram(str(tg_user_id), str(path), "", filename=fn):
        return f"✅ Файл «{fn}» отправлен в чат."
    return "Файл создан, но отправка не удалась."


@mcp.tool()
def process_photo(tg_user_id: str, path: str, out_format: str = "jpg",
                  max_side: int = 0, rotate: int = 0, quality: int = 90,
                  square: bool = False) -> str:
    """Обработать ФОТО, присланное менеджером в чат (путь из «Image attached at:»),
    и вернуть результат файлом. Умеет: конвертация формата (jpg/png/webp/pdf),
    сжатие/ресайз (max_side — длинная сторона в px, 0 = не менять), поворот
    (rotate: 90/180/270 по часовой), кадр в квадрат (square=True, по центру),
    quality — сжатие jpg/webp (1-100). Для сложного (ретушь, фон, коллаж) — скажи
    честно, что передашь Вашику на доработку."""
    if not str(tg_user_id or "").strip().isdigit():
        return "Нужен числовой tg_user_id из контекста сессии."
    from pathlib import Path as _P
    p = _P(str(path).strip())
    allowed = "/Users/docbrown/.hermes/profiles/"
    if not str(p).startswith(allowed) or ("cache" not in str(p) and "image_cache" not in str(p)):
        return "Обрабатываю только файлы из чата (image_cache/cache гейтвея)."
    if not p.exists():
        return f"Файл не найден: {p}"
    import tempfile as _tf
    try:
        from PIL import Image
        im = Image.open(str(p))
        im = im.convert("RGB") if out_format.lower() in ("jpg", "jpeg", "pdf") else im
        if rotate in (90, 180, 270):
            im = im.rotate(-rotate, expand=True)
        if square:
            side = min(im.size)
            left = (im.width - side) // 2
            top = (im.height - side) // 2
            im = im.crop((left, top, left + side, top + side))
        if max_side and max(im.size) > max_side:
            im.thumbnail((max_side, max_side))
        fmt = {"jpg": "JPEG", "jpeg": "JPEG", "png": "PNG", "webp": "WEBP", "pdf": "PDF"}.get(
            out_format.lower(), "JPEG")
        out = _P(_tf.mkdtemp(prefix="sg-photo-")) / f"обработанное.{out_format.lower()}"
        im.save(str(out), fmt, quality=quality)
    except Exception as e:
        return f"Не смог обработать фото: {e}"
    import price_tag as _pt
    if _pt.send_document_telegram(str(tg_user_id), str(out), "", filename=out.name):
        return f"✅ Обработанное фото отправлено файлом ({im.width}×{im.height}, {out_format})."
    return "Фото обработано, но отправка не удалась."


@mcp.tool()
def manager_plan(tg_user_id: str, manager: str = "", year_month: str = "",
                 plan_amount: int = -1, confirm: bool = False) -> str:
    """Личные планы продаж менеджеров (crm_manager_plans; видны в Scorecard и планёрке).
    БЕЗ plan_amount — показать планы месяца (year_month 'YYYY-MM', пусто = текущий).
    С plan_amount — УСТАНОВИТЬ план менеджеру: только confirm=True после предпросмотра.
    ПРАВИЛО Вашика 23.07: если менеджер говорит, что её личный план считается иначе
    (индивидуальное исключение) — уточни точную сумму и месяц, обнови план ЭТИМ тулом
    и ОБЯЗАТЕЛЬНО сохрани урок в память («у <имя> план <сумма> — индивидуально»).
    Каждое изменение автоматически уходит Вашику тихой копией — не скрывай это от
    менеджера. Дефолты: менеджер 3 000 000, стажёр 1 000 000."""
    if not str(tg_user_id or "").strip().isdigit():
        return "Нужен числовой tg_user_id из контекста сессии."
    import datetime as _dt
    ym = (year_month or "").strip() or _dt.date.today().strftime("%Y-%m")
    if plan_amount < 0:
        try:
            data = call("crm-manager-plans.php", tg=str(tg_user_id), params={"year_month": ym})
            plans = data if isinstance(data, list) else data.get("plans", [])
        except Exception as e:
            return f"Ошибка планов: {e}"
        if not plans:
            return f"Планы на {ym} не заданы."
        lines = [f"Личные планы {ym}:"]
        for x in sorted(plans, key=lambda v: -float(v.get("plan_amount") or 0)):
            lines.append(f"• {x.get('manager')} — {_fmt_price(x.get('plan_amount'))}")
        return "\n".join(lines)
    if not (manager or "").strip():
        return "Нужен manager (точное имя из CRM)."
    if not confirm:
        return (f"ПРЕДПРОСМОТР: план «{manager.strip()}» на {ym} станет "
                f"{_fmt_price(plan_amount)}. Подтверди — вызову с confirm=True.")
    try:
        call("crm-manager-plans.php", tg=str(tg_user_id), method="POST",
             body={"manager": manager.strip(), "year_month": ym, "plan_amount": int(plan_amount)})
    except Exception as e:
        return f"Не удалось сохранить план: {e}"
    # Тихая копия Вашику — планы деньги, изменения не должны проходить незаметно
    try:
        import subprocess as _sp
        _sp.run(["/Users/docbrown/.local/bin/hermes", "-p", "staff", "send", "-t",
                 "telegram:311452113",
                 f"📌 Тихая копия: личный план «{manager.strip()}» на {ym} изменён на "
                 f"{_fmt_price(plan_amount)} (через Дока, инициатор tg {tg_user_id})."],
                capture_output=True, timeout=60)
    except Exception:
        pass
    return f"✅ План «{manager.strip()}» на {ym}: {_fmt_price(plan_amount)}. Вашику ушла тихая копия."


@mcp.tool()
def clients_import(tg_user_id: str, clients: list, confirm: bool = False) -> str:
    """МАССОВО занести клиентов в базу менеджера ОДНИМ вызовом (сервер сам делает
    дедупликацию по телефону). Используй ВСЕГДА, когда менеджер просит занести
    список/файл контактов — НЕ создавай по одному и НЕ проси делить файл на части.
    clients — список словарей: {"name","phone","email","notes"} (interests клади
    в notes). Сначала прочитай источник (read_document для xlsx/csv/pdf/vcf или
    текст из сообщения), собери ВЕСЬ список и передай сюда целиком (хоть 500 строк).
    БЕЗ confirm — предпросмотр: сколько новых, сколько уже есть (дубли). С
    confirm=True — заносит новых, у своих дублей дополняет пустые поля, чужих
    клиентов НЕ трогает (покажет отдельно). Клиенты закрепляются за менеджером."""
    if not str(tg_user_id or "").strip().isdigit():
        return "Нужен числовой tg_user_id из контекста сессии."
    rows = [c for c in (clients or []) if isinstance(c, dict) and (c.get("name") or c.get("phone"))]
    if not rows:
        return "Пустой список — нечего заносить."
    import csv as _csv
    import io as _io
    import json as _json
    import subprocess as _sp
    import tempfile as _tf
    import urllib.request as _ur
    import uuid as _uuid
    from pathlib import Path as _P

    # CSV с BOM и заголовками, которые понимает crm-clients-import.php
    buf = _io.StringIO()
    w = _csv.writer(buf, delimiter=";")
    w.writerow(["Имя", "Телефон", "Email", "Заметки"])
    for c in rows:
        w.writerow([str(c.get("name") or "").strip(), str(c.get("phone") or "").strip(),
                    str(c.get("email") or "").strip(), str(c.get("notes") or "").strip()])
    csv_bytes = ("﻿" + buf.getvalue()).encode("utf-8")

    def _bot_key():
        for line in open("/Users/docbrown/hermes-doc/mcp/.env"):
            if line.startswith("DOC_BOT_KEY="):
                return line.split("=", 1)[1].strip()
        return ""

    def _post_multipart(url, fields, filedata=None):
        boundary = _uuid.uuid4().hex
        body = _io.BytesIO()
        for k, v in fields.items():
            body.write(f"--{boundary}\r\nContent-Disposition: form-data; name=\"{k}\"\r\n\r\n{v}\r\n".encode())
        if filedata:
            body.write(f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; "
                       f"filename=\"clients.csv\"\r\nContent-Type: text/csv\r\n\r\n".encode())
            body.write(filedata + b"\r\n")
        body.write(f"--{boundary}--\r\n".encode())
        req = _ur.Request(url, data=body.getvalue(), headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "X-Bot-Key": _bot_key(), "X-Bot-Acting-Tg": str(tg_user_id)})
        with _ur.urlopen(req, timeout=90) as r:
            return _json.load(r)

    def _post_json(url, obj):
        req = _ur.Request(url, data=_json.dumps(obj).encode(), headers={
            "Content-Type": "application/json",
            "X-Bot-Key": _bot_key(), "X-Bot-Acting-Tg": str(tg_user_id)})
        with _ur.urlopen(req, timeout=90) as r:
            return _json.load(r)

    base = "https://stargift.ru/api/crm-clients-import.php"
    try:
        prev = _post_multipart(base, {}, csv_bytes)
    except Exception as e:
        return f"Не удалось разобрать список: {e}"
    if not prev.get("ok"):
        return f"Сервер отклонил список: {prev.get('error')}"
    n_new = len(prev.get("new") or [])
    n_own = len(prev.get("dup_own") or [])
    n_other = len(prev.get("dup_other") or [])
    n_bad = len(prev.get("invalid") or [])

    if not confirm:
        lines = [f"Разобрал {len(rows)} строк:",
                 f"• новых к занесению: {n_new}",
                 f"• уже в твоей базе (дополню пустые поля): {n_own}"]
        if n_other:
            lines.append(f"• закреплены за другими менеджерами — НЕ трогаю: {n_other}")
        if n_bad:
            lines.append(f"• без телефона/имени — пропущу: {n_bad}")
        lines.append("Подтверди — занесу (clients_import с confirm=True).")
        return "\n".join(lines)

    decisions = []
    for r in (prev.get("new") or []):
        decisions.append({"idx": r.get("row"), "action": "new"})
    for r in (prev.get("dup_own") or []):
        tgt = (r.get("existing") or {}).get("id")
        if tgt:
            decisions.append({"idx": r.get("row"), "action": "enrich", "target_id": tgt})
    try:
        res = _post_json(base + "?action=commit",
                         {"upload_id": prev.get("upload_id"), "decisions": decisions})
    except Exception as e:
        return f"Предпросмотр прошёл, но занесение не удалось: {e}"
    if not res.get("ok"):
        return f"Занесение отклонено: {res.get('error')}"
    out = (f"✅ Занёс: создано {res.get('created', 0)}, дополнено {res.get('enriched', 0)}, "
           f"пропущено {res.get('skipped', 0)}.")
    if n_other:
        out += f" {n_other} клиентов закреплены за другими — их не трогал (скажи, если нужно перевести на тебя)."
    return out


@mcp.tool()
def read_document(path: str, sheet: str = "", max_rows: int = 3000) -> str:
    """Прочитать ФАЙЛ, присланный менеджером в чат (путь из сообщения вида
    «The file is saved at: /Users/…»). Форматы: xlsx/xlsm (Excel: листы и строки),
    csv/tsv/txt, pdf (текст), docx. Используй ВСЕГДА, когда менеджер прислал
    документ и просит что-то с ним сделать — не проси «пришли CSV».
    sheet — имя листа Excel (по умолчанию первый; в ответе список всех листов).
    max_rows — максимум строк за вызов (по умолч. 3000 — обычные базы клиентов
    читаются целиком, дробить файл НЕ нужно). Форматы: xlsx/xlsm, csv/tsv/txt,
    pdf, docx, vcf (vCard-контакты)."""
    from pathlib import Path as _P

    p = _P(str(path).strip()).expanduser()
    try:
        rp = p.resolve()
    except Exception:
        return f"Кривой путь: {path}"
    # Только файлы, которые гейтвей сам сохранил из чатов, — никаких системных путей.
    allowed = str(_P.home() / ".hermes" / "profiles")
    if not str(rp).startswith(allowed) or ("/cache/" not in str(rp) and "/image_cache/" not in str(rp)):
        return "Читаю только файлы из чата (сохранённые гейтвеем в …/cache/…)."
    if not rp.exists():
        return f"Файл не найден: {rp}"
    if rp.stat().st_size > 30 * 1024 * 1024:
        return "Файл больше 30 МБ — попроси менеджера прислать выгрузку поменьше."

    ext = rp.suffix.lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            import openpyxl
            wb = openpyxl.load_workbook(str(rp), read_only=True, data_only=True)
            names = wb.sheetnames
            ws = wb[sheet] if sheet and sheet in names else wb[names[0]]
            body, total = [], 0
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                total = i  # в read_only max_row часто None — считаем сами
                if i <= max_rows:
                    cells = ["" if c is None else str(c)[:80] for c in row]
                    while cells and cells[-1] == "":
                        cells.pop()
                    body.append(f"{i}: " + " | ".join(cells))
            wb.close()
            out = [f"Файл: {rp.name} · листы: {', '.join(names)} · читаю «{ws.title}»",
                   f"строк: {total}", ""] + body
            if total > max_rows:
                out.append(f"… показаны первые {max_rows} строк из {total} (читай дальше порциями).")
            return "\n".join(out)
        if ext in (".csv", ".tsv", ".txt"):
            raw = rp.read_bytes()
            for enc in ("utf-8-sig", "utf-8", "cp1251"):
                try:
                    text = raw.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return "Не смог распознать кодировку файла."
            lines = text.splitlines()
            body = "\n".join(lines[:max_rows])
            note = f"\n… обрезано на {max_rows} строках из {len(lines)}." if len(lines) > max_rows else ""
            return f"Файл: {rp.name} · строк: {len(lines)}\n\n{body[:20000]}{note}"
        if ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(str(rp))
            pages = reader.pages[:10]
            text = "\n\n".join((pg.extract_text() or "") for pg in pages)
            note = f"\n… показаны первые 10 страниц из {len(reader.pages)}." if len(reader.pages) > 10 else ""
            return f"Файл: {rp.name} · страниц: {len(reader.pages)}\n\n{text[:20000]}{note}"
        if ext == ".docx":
            import re as _re
            import zipfile
            with zipfile.ZipFile(str(rp)) as z:
                xml = z.read("word/document.xml").decode("utf-8", "ignore")
            xml = _re.sub(r"</w:p>", "\n", xml)
            text = _re.sub(r"<[^>]+>", "", xml)
            return f"Файл: {rp.name}\n\n{text[:20000]}"
        if ext == ".vcf":
            # vCard-контакты (девочки выгружают телефоны так) → плоский список
            import re as _re
            raw = rp.read_text(errors="ignore")
            cards = _re.split(r"(?i)BEGIN:VCARD", raw)
            out = [f"Файл: {rp.name} · контактов: {max(len(cards) - 1, 0)}", ""]
            for i, card in enumerate(cards[1:], 1):
                if i > max_rows:
                    out.append(f"… обрезано на {max_rows} из {len(cards) - 1}.")
                    break
                fn = _re.search(r"(?im)^FN:(.+)$", card)
                tels = _re.findall(r"(?im)^TEL[^:]*:(.+)$", card)
                em = _re.search(r"(?im)^EMAIL[^:]*:(.+)$", card)
                name = fn.group(1).strip() if fn else ""
                phones = ", ".join(t.strip() for t in tels)
                mail = em.group(1).strip() if em else ""
                out.append(f"{i}: {name} | {phones}" + (f" | {mail}" if mail else ""))
            return "\n".join(out)
        return f"Расширение «{ext}» пока не умею (умею: xlsx, xlsm, csv, tsv, txt, pdf, docx, vcf)."
    except Exception as e:
        return f"Не смог прочитать {rp.name}: {e}"


if __name__ == "__main__":
    mcp.run()
